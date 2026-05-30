"""FastAPI endpoint tests — exercises every route in main.py.

Uses the real checks.json and real ideal files from ideals/ so tests
reflect what ships to end users. A minimal in-memory actual file is
built for each module to avoid needing real client data on disk.
"""

import io
import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from fastapi.testclient import TestClient

from main import app

client = TestClient(app)


# ── helpers ───────────────────────────────────────────────────────────────────


def make_xlsx_bytes(sheet_name: str, rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_csv_bytes(content: str) -> bytes:
    return content.encode()


# Minimal actual files for each stored check
INV_ACTUAL = make_xlsx_bytes(
    "INV_ORGANIZATION_PARAMETER",
    [
        [
            "Organization",
            "InventoryFlag",
            "MfgPlantFlag",
            "AllowItemSubstitutions",
            "AllowNegOnhandCcTxns",
            "EamEnabledFlag",
            "FifoOrigRcptDateFlag",
            "FABookTypeCode",
            "BusinessUnitName",
            "OrganizationCode",
            "MasterOrgCode",
            "ProfitCenterBuName",
        ],
        [
            "Mumbai Plant",
            "Y",
            "Y",
            "Y",
            "N",
            "Y",
            "Y",
            "N",
            "Mumbai BU",
            2,
            0,
            "Mumbai BU",
        ],
        ["Delhi HQ", "Y", "Y", "N", "N", "Y", "Y", "N", "Delhi BU", 50, 0, "Delhi BU"],
    ],
)

RCV_ACTUAL = make_xlsx_bytes(
    "RCV_OPTIONS",
    [
        [
            "Organization",
            "AllowCascadeReceipts",
            "AllowSubstitutions",
            "AllowUnorderedReceipts",
            "ReceivingRoutingId",
        ],
        ["Mumbai Plant", "Y", "N", "N", "1"],
        ["Delhi HQ", "Y", "N", "N", "1"],
    ],
)

INV_TOL_ACTUAL = make_xlsx_bytes(
    "AP_TOLERANCE_TEMPLATE",
    [
        [
            "BU Name",
            "MaxQuantityOrdered",
            "MaxQuantityReceived",
            "MaxAmountOrdered",
            "MaxAmountReceived",
        ],
        ["Mumbai Plant", "5.0", "5.0", "500.0", "500.0"],
        ["Delhi HQ", "5.0", "5.0", "500.0", "500.0"],
    ],
)

PAY_PROC_ACTUAL = make_xlsx_bytes(
    "AP_FINANCIAL_SYS_PARAM",
    [
        [
            "BU Name",
            "PositivePayRequired",
            "AutoCalculateTaxFlag",
            "PrepaymentPayGroup",
            "DefaultExchangeRateType",
        ],
        ["Mumbai Plant", "Y", "Y", "PREPAY", "Corporate"],
        ["Delhi HQ", "N", "Y", "PREPAY", "Corporate"],
    ],
)

INVOICE_OPT_ACTUAL = make_csv_bytes(
    "BU,InvoiceMatchOption,HoldUnmatchedInvoices,TaxCalculationLevel\n"
    "Mumbai Plant,PO,N,Line\n"
    "Delhi HQ,PO,N,Line\n"
)


# ── GET /api/checks ───────────────────────────────────────────────────────────


class TestListChecks:
    def test_returns_200(self):
        r = client.get("/api/checks")
        assert r.status_code == 200

    def test_returns_list(self):
        r = client.get("/api/checks")
        data = r.json()
        assert isinstance(data, list)
        assert len(data) >= 1

    def test_each_check_has_required_fields(self):
        r = client.get("/api/checks")
        for check in r.json():
            assert "id" in check
            assert "name" in check
            assert "default_sheet" in check

    def test_five_checks_present(self):
        r = client.get("/api/checks")
        ids = {c["id"] for c in r.json()}
        assert "inv_organization_parameters" in ids
        assert "receiving_parameters" in ids
        assert "invoice_options" in ids
        assert "invoice_tolerances" in ids
        assert "common_pay_proc_options" in ids


# ── POST /api/audit ───────────────────────────────────────────────────────────


class TestRunAudit:
    def test_inv_org_parameters_returns_excel(self):
        r = client.post(
            "/api/audit?check_id=inv_organization_parameters",
            files={
                "actual_file": (
                    "inv.xlsx",
                    INV_ACTUAL,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert r.status_code == 200
        content_type = r.headers.get("content-type", "")
        # dev mode returns spreadsheet; frozen returns json
        assert "spreadsheet" in content_type or "json" in content_type

    def test_receiving_parameters_xlsx(self):
        r = client.post(
            "/api/audit?check_id=receiving_parameters",
            files={
                "actual_file": (
                    "rcv.xlsx",
                    RCV_ACTUAL,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert r.status_code == 200

    def test_invoice_tolerances_xlsx(self):
        r = client.post(
            "/api/audit?check_id=invoice_tolerances",
            files={
                "actual_file": (
                    "tol.xlsx",
                    INV_TOL_ACTUAL,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert r.status_code == 200

    def test_common_pay_proc_options_xlsx(self):
        r = client.post(
            "/api/audit?check_id=common_pay_proc_options",
            files={
                "actual_file": (
                    "pay.xlsx",
                    PAY_PROC_ACTUAL,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert r.status_code == 200

    def test_invoice_options_csv(self):
        r = client.post(
            "/api/audit?check_id=invoice_options",
            files={"actual_file": ("inv_opt.csv", INVOICE_OPT_ACTUAL, "text/csv")},
        )
        assert r.status_code == 200

    def test_missing_check_id_and_ideal_file_returns_400(self):
        r = client.post(
            "/api/audit",
            files={
                "actual_file": (
                    "inv.xlsx",
                    INV_ACTUAL,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert r.status_code == 400

    def test_unknown_check_id_returns_404(self):
        r = client.post(
            "/api/audit?check_id=nonexistent_check",
            files={
                "actual_file": (
                    "inv.xlsx",
                    INV_ACTUAL,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert r.status_code == 404
        assert "nonexistent_check" in r.json()["detail"]

    def test_wrong_sheet_name_returns_400(self):
        r = client.post(
            "/api/audit?check_id=inv_organization_parameters&sheet_name=WRONG_SHEET",
            files={
                "actual_file": (
                    "inv.xlsx",
                    INV_ACTUAL,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert r.status_code == 400
        assert "WRONG_SHEET" in r.json()["detail"]

    def test_custom_fuzzy_threshold(self):
        r = client.post(
            "/api/audit?check_id=inv_organization_parameters&fuzzy_threshold=90",
            files={
                "actual_file": (
                    "inv.xlsx",
                    INV_ACTUAL,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert r.status_code == 200

    def test_threshold_below_50_rejected(self):
        r = client.post(
            "/api/audit?check_id=inv_organization_parameters&fuzzy_threshold=30",
            files={
                "actual_file": (
                    "inv.xlsx",
                    INV_ACTUAL,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert r.status_code == 422

    def test_threshold_above_100_rejected(self):
        r = client.post(
            "/api/audit?check_id=inv_organization_parameters&fuzzy_threshold=101",
            files={
                "actual_file": (
                    "inv.xlsx",
                    INV_ACTUAL,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert r.status_code == 422

    def test_ideal_file_override(self):
        """Uploading an ideal_file directly bypasses the stored check."""
        ideal_csv = b"Config Name,Ideal Value\nInventoryFlag,Y\n"
        r = client.post(
            "/api/audit",
            files={
                "actual_file": (
                    "inv.xlsx",
                    INV_ACTUAL,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
                "ideal_file": ("ideal.csv", ideal_csv, "text/csv"),
            },
        )
        assert r.status_code == 200

    def test_report_is_valid_xlsx_in_dev_mode(self):
        """POST /api/audit returns a token; GET /api/download/{token} returns xlsx."""
        r = client.post(
            "/api/audit?check_id=inv_organization_parameters",
            files={
                "actual_file": (
                    "inv.xlsx",
                    INV_ACTUAL,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert r.status_code == 200
        token = r.json()["token"]
        dl = client.get(f"/api/download/{token}")
        assert dl.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(dl.content))
        assert "Controls in place" in wb.sheetnames
        assert "Control gaps" in wb.sheetnames
        assert "Audit Summary" in wb.sheetnames


# ── GET /api/admin/checks/{id}/download ──────────────────────────────────────


class TestAdminDownload:
    def test_download_inv_org_parameters(self):
        r = client.get("/api/admin/checks/inv_organization_parameters/download")
        assert r.status_code == 200
        # v1 returns a token; redeem it for the xlsx bytes
        token = r.json()["token"]
        dl = client.get(f"/api/download/{token}")
        assert dl.status_code == 200
        assert "spreadsheet" in dl.headers.get("content-type", "")

    def test_download_all_checks(self):
        checks = client.get("/api/checks").json()
        for check in checks:
            r = client.get(f"/api/admin/checks/{check['id']}/download")
            assert r.status_code == 200, f"download failed for {check['id']}"
            token = r.json()["token"]
            dl = client.get(f"/api/download/{token}")
            assert dl.status_code == 200, f"token download failed for {check['id']}"

    def test_download_unknown_check_returns_404(self):
        r = client.get("/api/admin/checks/nonexistent/download")
        assert r.status_code == 404


# ── POST /api/admin/checks/{id}/upload ───────────────────────────────────────


class TestAdminUpload:
    def test_upload_replaces_ideal_file(self):
        # Download current, re-upload the same bytes — idempotent
        download = client.get("/api/admin/checks/inv_organization_parameters/download")
        original_bytes = download.content

        r = client.post(
            "/api/admin/checks/inv_organization_parameters/upload",
            files={
                "file": (
                    "inv_org.xlsx",
                    original_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert r.status_code == 200
        assert r.json()["status"] == "ok"

    def test_upload_unknown_check_returns_404(self):
        r = client.post(
            "/api/admin/checks/nonexistent/upload",
            files={"file": ("x.xlsx", b"data", "application/octet-stream")},
        )
        assert r.status_code == 404
