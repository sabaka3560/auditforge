"""Oracle IT audit scenario tests.

Each test proves a specific audit outcome: a real misconfiguration is detected,
a compliant config is confirmed, or a capture field is extracted correctly.
Tests are written from the auditor's perspective — what should appear in which
sheet of the report — not from the code's perspective.

Five Oracle Fusion modules covered:
  1. INV Organization Parameters
  2. Receiving (RCV) Parameters
  3. AP Invoice Options
  4. AP Invoice Tolerances
  5. AP Common Pay Proc Options
"""

import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pandas as pd

from engine import build_mapping, build_report, compare
from engine.comparator import compare
from engine.matcher import build_mapping


# ── shared helpers ─────────────────────────────────────────────────────────────


BUS_INDIA = ["Mumbai Plant", "Delhi HQ", "Bangalore IT", "Chennai Plant", "Pune SEZ"]
BUS_GLOBAL = [
    "Singapore Regional",
    "UAE Free Zone",
    "London EMEA",
    "New York Americas",
    "Tokyo APAC",
]
ALL_BUS = BUS_INDIA + BUS_GLOBAL


def actual(bus: list[str], configs: dict) -> pd.DataFrame:
    return pd.DataFrame({"BU_NAME": bus, **configs})


def ideal(rows: list[tuple[str, str, str]]) -> pd.DataFrame:
    """rows: (config_name, ideal_value, options)"""
    df = pd.DataFrame(rows, columns=["config_name", "ideal_value", "options"])
    df["options"] = df["options"].fillna("")
    return df


def run(bus, configs, ideal_rows):
    actual_df = actual(bus, configs)
    ideal_df = ideal(ideal_rows)
    mapping = build_mapping(ideal_df, list(actual_df.columns))
    cip, gaps, extra = compare(actual_df, mapping)
    return cip, gaps, extra, mapping


def gap_keys(gaps):
    return {(r.bu_name, r.config_name) for r in gaps}


def cip_keys(cip):
    return {(r.bu_name, r.config_name) for r in cip}


def captured_keys(extra):
    return {(r.bu_name, r.config_name) for r in extra}


# ══════════════════════════════════════════════════════════════════════════════
# MODULE 1 — INV Organization Parameters
# ══════════════════════════════════════════════════════════════════════════════


INV_IDEAL = [
    ("AllowItemSubstitutionsFlag", "Y", "Y, N"),
    ("AllowNegOnhandCcTxns", "N", "Y, N"),
    ("EamEnabledFlag", "Y", "Y, N"),
    ("FifoOrigRcptDateFlag", "Y", "Y, N"),
    ("InventoryFlag", "Y", "Y, N"),
    ("MfgPlantFlag", "Y", "Y, N"),
    ("FaBookTypeCode", "N", "Y, N"),
    # Capture-only
    ("BusinessUnitId", "Capture the value defined", ""),
    ("OrganizationId", "Capture the value defined", ""),
    ("MasterOrganizationId", "Capture the value defined", ""),
    ("ProfitCenterBuId", "Capture the value defined", ""),
]


class TestINVOrganizationParameters:
    def test_negative_inventory_allowed_is_flagged(self):
        """AllowNegOnhandCcTxns = Y when ideal is N → control gap for that BU."""
        cip, gaps, _, _ = run(
            BUS_INDIA,
            {
                "AllowNegOnhandCcTxns": ["N", "Y", "N", "N", "N"],  # Delhi HQ = Y ← gap
                "AllowItemSubstitutions": ["Y"] * 5,
                "EamEnabledFlag": ["Y"] * 5,
                "FifoOrigRcptDateFlag": ["Y"] * 5,
                "InventoryFlag": ["Y"] * 5,
                "MfgPlantFlag": ["Y"] * 5,
                "FABookTypeCode": ["N"] * 5,
                "BusinessUnitName": ["BU"] * 5,
                "OrganizationCode": [1, 2, 3, 4, 5],
                "MasterOrgCode": [0] * 5,
                "ProfitCenterBuName": ["BU"] * 5,
            },
            INV_IDEAL,
        )
        assert ("Delhi HQ", "AllowNegOnhandCcTxns") in gap_keys(gaps)
        assert ("Mumbai Plant", "AllowNegOnhandCcTxns") not in gap_keys(gaps)

    def test_inventory_disabled_is_critical_gap(self):
        """InventoryFlag = N means inventory is turned off — always a gap."""
        cip, gaps, _, _ = run(
            BUS_INDIA,
            {
                "InventoryFlag": ["Y", "N", "Y", "Y", "Y"],  # Delhi HQ off ← gap
                "AllowNegOnhandCcTxns": ["N"] * 5,
                "AllowItemSubstitutions": ["Y"] * 5,
                "EamEnabledFlag": ["Y"] * 5,
                "FifoOrigRcptDateFlag": ["Y"] * 5,
                "MfgPlantFlag": ["Y"] * 5,
                "FABookTypeCode": ["N"] * 5,
                "BusinessUnitName": ["BU"] * 5,
                "OrganizationCode": [1, 2, 3, 4, 5],
                "MasterOrgCode": [0] * 5,
                "ProfitCenterBuName": ["BU"] * 5,
            },
            INV_IDEAL,
        )
        assert ("Delhi HQ", "InventoryFlag") in gap_keys(gaps)

    def test_null_eam_flag_is_gap_not_cip(self):
        """Null EamEnabledFlag = missing config = control gap, not in place."""
        cip, gaps, _, _ = run(
            BUS_INDIA,
            {
                "EamEnabledFlag": ["Y", None, "Y", None, "Y"],
                "AllowNegOnhandCcTxns": ["N"] * 5,
                "AllowItemSubstitutions": ["Y"] * 5,
                "FifoOrigRcptDateFlag": ["Y"] * 5,
                "InventoryFlag": ["Y"] * 5,
                "MfgPlantFlag": ["Y"] * 5,
                "FABookTypeCode": ["N"] * 5,
                "BusinessUnitName": ["BU"] * 5,
                "OrganizationCode": [1, 2, 3, 4, 5],
                "MasterOrgCode": [0] * 5,
                "ProfitCenterBuName": ["BU"] * 5,
            },
            INV_IDEAL,
        )
        null_bu_keys = gap_keys(gaps)
        assert ("Delhi HQ", "EamEnabledFlag") in null_bu_keys
        assert (
            "Singapore Regional" if len(BUS_INDIA) > 3 else "Chennai Plant",
            "EamEnabledFlag",
        ) in null_bu_keys or ("Chennai Plant", "EamEnabledFlag") in null_bu_keys

    def test_org_code_is_captured_not_compared(self):
        """OrganizationCode maps to OrganizationId (capture) — must appear in extra, never gaps."""
        cip, gaps, extra, _ = run(
            BUS_INDIA,
            {
                "OrganizationCode": [2, 50, 51, 200, 300],
                "InventoryFlag": ["Y"] * 5,
                "AllowNegOnhandCcTxns": ["N"] * 5,
                "AllowItemSubstitutions": ["Y"] * 5,
                "EamEnabledFlag": ["Y"] * 5,
                "FifoOrigRcptDateFlag": ["Y"] * 5,
                "MfgPlantFlag": ["Y"] * 5,
                "FABookTypeCode": ["N"] * 5,
                "BusinessUnitName": ["BU"] * 5,
                "MasterOrgCode": [0] * 5,
                "ProfitCenterBuName": ["BU"] * 5,
            },
            INV_IDEAL,
        )
        captured = {r.config_name for r in extra}
        assert "OrganizationId" in captured
        assert all(r.config_name != "OrganizationId" for r in gaps)

    def test_org_code_values_are_preserved_in_capture(self):
        cip, gaps, extra, _ = run(
            ["Mumbai Plant", "Delhi HQ"],
            {
                "OrganizationCode": [2, 50],
                "InventoryFlag": ["Y", "Y"],
                "AllowNegOnhandCcTxns": ["N", "N"],
                "AllowItemSubstitutions": ["Y", "Y"],
                "EamEnabledFlag": ["Y", "Y"],
                "FifoOrigRcptDateFlag": ["Y", "Y"],
                "MfgPlantFlag": ["Y", "Y"],
                "FABookTypeCode": ["N", "N"],
                "BusinessUnitName": ["Mumbai BU", "Delhi BU"],
                "MasterOrgCode": [0, 0],
                "ProfitCenterBuName": ["Mumbai BU", "Delhi BU"],
            },
            INV_IDEAL,
        )
        org_captures = {
            r.bu_name: r.actual_value
            for r in extra
            if r.config_name == "OrganizationId"
        }
        assert org_captures["Mumbai Plant"] == "2"
        assert org_captures["Delhi HQ"] == "50"

    def test_boolean_case_insensitive_y_passes(self):
        """Oracle exports sometimes use lowercase 'y' — must not be flagged."""
        cip, gaps, _, _ = run(
            ["Mumbai Plant"],
            {
                "InventoryFlag": ["y"],
                "AllowNegOnhandCcTxns": ["n"],
                "AllowItemSubstitutions": ["yes"],
                "EamEnabledFlag": ["Y"],
                "FifoOrigRcptDateFlag": ["Y"],
                "MfgPlantFlag": ["Y"],
                "FABookTypeCode": ["N"],
                "BusinessUnitName": ["BU"],
                "OrganizationCode": [1],
                "MasterOrgCode": [0],
                "ProfitCenterBuName": ["BU"],
            },
            INV_IDEAL,
        )
        inv_cip = [r for r in cip if r.config_name == "InventoryFlag"]
        inv_gaps = [r for r in gaps if r.config_name == "InventoryFlag"]
        assert len(inv_cip) == 1
        assert len(inv_gaps) == 0

    def test_all_compliant_bu_has_zero_gaps(self):
        values = {
            "InventoryFlag": ["Y"] * 5,
            "AllowNegOnhandCcTxns": ["N"] * 5,
            "AllowItemSubstitutions": ["Y"] * 5,
            "EamEnabledFlag": ["Y"] * 5,
            "FifoOrigRcptDateFlag": ["Y"] * 5,
            "MfgPlantFlag": ["Y"] * 5,
            "FABookTypeCode": ["N"] * 5,
            "BusinessUnitName": ["BU"] * 5,
            "OrganizationCode": [1, 2, 3, 4, 5],
            "MasterOrgCode": [0] * 5,
            "ProfitCenterBuName": ["BU"] * 5,
        }
        cip, gaps, _, _ = run(BUS_INDIA, values, INV_IDEAL)
        assert len(gaps) == 0


# ══════════════════════════════════════════════════════════════════════════════
# MODULE 2 — Receiving (RCV) Parameters
# ══════════════════════════════════════════════════════════════════════════════


RCV_IDEAL = [
    ("AllowCascadeReceipts", "Y", "Y, N"),
    ("AllowSubstitutions", "N", "Y, N"),
    ("AllowUnorderedReceipts", "N", "Y, N"),
    ("ReceivingRoutingId", "1", "1-Direct, 2-Standard, 3-Inspection"),
    ("EnforceShipToLocation", "N", "Y, N"),
    # Capture
    ("OrganizationCode", "Capture", ""),
]


class TestReceivingParameters:
    def test_substitutions_allowed_is_gap(self):
        """AllowSubstitutions=Y means unapproved items can be received — audit gap."""
        cip, gaps, _, _ = run(
            BUS_INDIA,
            {
                "AllowCascadeReceipts": ["Y"] * 5,
                "AllowSubstitutions": ["N", "Y", "N", "N", "N"],  # Delhi HQ = Y ← gap
                "AllowUnorderedReceipts": ["N"] * 5,
                "ReceivingRoutingId": ["1"] * 5,
                "EnforceShipToLocation": ["N"] * 5,
                "OrganizationCode": [1, 2, 3, 4, 5],
            },
            RCV_IDEAL,
        )
        assert ("Delhi HQ", "AllowSubstitutions") in gap_keys(gaps)
        assert ("Mumbai Plant", "AllowSubstitutions") not in gap_keys(gaps)

    def test_unordered_receipts_allowed_is_gap(self):
        """AllowUnorderedReceipts=Y is a segregation of duties risk."""
        cip, gaps, _, _ = run(
            BUS_INDIA,
            {
                "AllowCascadeReceipts": ["Y"] * 5,
                "AllowSubstitutions": ["N"] * 5,
                "AllowUnorderedReceipts": [
                    "Y",
                    "N",
                    "Y",
                    "N",
                    "N",
                ],  # Mumbai, Bangalore ← gap
                "ReceivingRoutingId": ["1"] * 5,
                "EnforceShipToLocation": ["N"] * 5,
                "OrganizationCode": [1, 2, 3, 4, 5],
            },
            RCV_IDEAL,
        )
        assert ("Mumbai Plant", "AllowUnorderedReceipts") in gap_keys(gaps)
        assert ("Bangalore IT", "AllowUnorderedReceipts") in gap_keys(gaps)
        assert ("Delhi HQ", "AllowUnorderedReceipts") not in gap_keys(gaps)

    def test_wrong_routing_id_is_gap(self):
        """ReceivingRoutingId=2 (Standard) or 3 (Inspection) instead of 1 (Direct) = gap."""
        cip, gaps, _, _ = run(
            ["Mumbai Plant", "Delhi HQ", "Bangalore IT"],
            {
                "AllowCascadeReceipts": ["Y"] * 3,
                "AllowSubstitutions": ["N"] * 3,
                "AllowUnorderedReceipts": ["N"] * 3,
                "ReceivingRoutingId": [
                    "1",
                    "2",
                    "3",
                ],  # Delhi=Standard, Bangalore=Inspection ← gaps
                "EnforceShipToLocation": ["N"] * 3,
                "OrganizationCode": [1, 2, 3],
            },
            RCV_IDEAL,
        )
        assert ("Delhi HQ", "ReceivingRoutingId") in gap_keys(gaps)
        assert ("Bangalore IT", "ReceivingRoutingId") in gap_keys(gaps)
        assert ("Mumbai Plant", "ReceivingRoutingId") not in gap_keys(gaps)

    def test_routing_id_float_normalized(self):
        """Oracle sometimes exports routing ID as '1.0' — must equal ideal '1'."""
        cip, gaps, _, _ = run(
            ["Mumbai Plant"],
            {
                "AllowCascadeReceipts": ["Y"],
                "AllowSubstitutions": ["N"],
                "AllowUnorderedReceipts": ["N"],
                "ReceivingRoutingId": [
                    "1.0"
                ],  # Oracle float export ← should NOT be a gap
                "EnforceShipToLocation": ["N"],
                "OrganizationCode": [1],
            },
            RCV_IDEAL,
        )
        assert ("Mumbai Plant", "ReceivingRoutingId") not in gap_keys(gaps)

    def test_org_code_is_captured_in_rcv(self):
        cip, gaps, extra, _ = run(
            ["Mumbai Plant"],
            {
                "AllowCascadeReceipts": ["Y"],
                "AllowSubstitutions": ["N"],
                "AllowUnorderedReceipts": ["N"],
                "ReceivingRoutingId": ["1"],
                "EnforceShipToLocation": ["N"],
                "OrganizationCode": [99],
            },
            RCV_IDEAL,
        )
        captured = {r.config_name: r.actual_value for r in extra}
        assert "OrganizationCode" in captured
        assert captured["OrganizationCode"] == "99"


# ══════════════════════════════════════════════════════════════════════════════
# MODULE 3 — AP Invoice Options
# ══════════════════════════════════════════════════════════════════════════════


INV_OPT_IDEAL = [
    ("InvoiceMatchOption", "PO", "PO, Receipt, Order"),
    ("HoldUnmatchedInvoices", "N", "Y, N"),
    ("TaxCalculationLevel", "Line", "Line, Header"),
    ("AutoCalculateTaxFlag", "Y", "Y, N"),
    ("PaymentTerms", "Date", ""),  # capture-only (type descriptor)
    ("CurrencyConversionRateType", "Date", ""),  # capture-only
]


class TestAPInvoiceOptions:
    def test_receipt_match_is_gap(self):
        """InvoiceMatchOption=Receipt means invoices not matched to PO — audit gap."""
        cip, gaps, _, _ = run(
            BUS_INDIA,
            {
                "InvoiceMatchOption": ["PO", "Receipt", "PO", "PO", "Order"],
                "HoldUnmatchedInvoices": ["N"] * 5,
                "TaxCalculationLevel": ["Line"] * 5,
                "AutoCalculateTaxFlag": ["Y"] * 5,
                "PaymentTerms": ["Net30"] * 5,
                "CurrencyConversionRateType": ["Corporate"] * 5,
            },
            INV_OPT_IDEAL,
        )
        assert ("Delhi HQ", "InvoiceMatchOption") in gap_keys(gaps)
        assert ("Pune SEZ", "InvoiceMatchOption") in gap_keys(gaps)  # Order ≠ PO
        assert ("Mumbai Plant", "InvoiceMatchOption") not in gap_keys(gaps)

    def test_options_column_appears_in_gap_rows(self):
        """The valid options string must be carried through to gap rows for auditor context."""
        cip, gaps, _, _ = run(
            ["Delhi HQ"],
            {
                "InvoiceMatchOption": ["Receipt"],
                "HoldUnmatchedInvoices": ["N"],
                "TaxCalculationLevel": ["Line"],
                "AutoCalculateTaxFlag": ["Y"],
                "PaymentTerms": ["Net30"],
                "CurrencyConversionRateType": ["Corporate"],
            },
            INV_OPT_IDEAL,
        )
        match_gap = next(r for r in gaps if r.config_name == "InvoiceMatchOption")
        assert "PO" in match_gap.options
        assert "Receipt" in match_gap.options

    def test_hold_unmatched_enabled_is_gap(self):
        """HoldUnmatchedInvoices=Y means invoices are being held — unusual, flag it."""
        cip, gaps, _, _ = run(
            ["Mumbai Plant"],
            {
                "InvoiceMatchOption": ["PO"],
                "HoldUnmatchedInvoices": ["Y"],  # ← gap (ideal = N)
                "TaxCalculationLevel": ["Line"],
                "AutoCalculateTaxFlag": ["Y"],
                "PaymentTerms": ["Net30"],
                "CurrencyConversionRateType": ["Corporate"],
            },
            INV_OPT_IDEAL,
        )
        assert ("Mumbai Plant", "HoldUnmatchedInvoices") in gap_keys(gaps)

    def test_payment_terms_is_captured_not_compared(self):
        """PaymentTerms ideal='Date' (type descriptor) → capture, no comparison."""
        cip, gaps, extra, _ = run(
            ["Mumbai Plant"],
            {
                "InvoiceMatchOption": ["PO"],
                "HoldUnmatchedInvoices": ["N"],
                "TaxCalculationLevel": ["Line"],
                "AutoCalculateTaxFlag": ["Y"],
                "PaymentTerms": ["Net30"],
                "CurrencyConversionRateType": ["Corporate"],
            },
            INV_OPT_IDEAL,
        )
        assert all(r.config_name != "PaymentTerms" for r in gaps)
        assert any(r.config_name == "PaymentTerms" for r in extra)

    def test_header_tax_level_is_gap(self):
        """TaxCalculationLevel=Header instead of Line = incorrect setup."""
        cip, gaps, _, _ = run(
            ["Mumbai Plant"],
            {
                "InvoiceMatchOption": ["PO"],
                "HoldUnmatchedInvoices": ["N"],
                "TaxCalculationLevel": ["Header"],  # ← gap
                "AutoCalculateTaxFlag": ["Y"],
                "PaymentTerms": ["Net30"],
                "CurrencyConversionRateType": ["Corporate"],
            },
            INV_OPT_IDEAL,
        )
        assert ("Mumbai Plant", "TaxCalculationLevel") in gap_keys(gaps)


# ══════════════════════════════════════════════════════════════════════════════
# MODULE 4 — AP Invoice Tolerances
# ══════════════════════════════════════════════════════════════════════════════


TOL_IDEAL = [
    ("MaxQuantityOrdered", "5", ""),
    ("MaxQuantityReceived", "5", ""),
    ("MaxAmountOrdered", "10", ""),
    ("MaxAmountReceived", "10", ""),
    ("TaxAmountRange", "Capture the value defined", ""),
]


class TestAPInvoiceTolerances:
    def test_tolerance_above_limit_is_gap(self):
        """MaxQuantityOrdered=10 when ideal=5 means looser tolerance — control gap."""
        cip, gaps, _, _ = run(
            BUS_INDIA,
            {
                "MaxQuantityOrdered": ["5", "10", "5", "5", "5"],  # Delhi HQ = 10 ← gap
                "MaxQuantityReceived": ["5"] * 5,
                "MaxAmountOrdered": ["10"] * 5,
                "MaxAmountReceived": ["10"] * 5,
                "TaxAmountRange": ["500"] * 5,
            },
            TOL_IDEAL,
        )
        assert ("Delhi HQ", "MaxQuantityOrdered") in gap_keys(gaps)
        assert ("Mumbai Plant", "MaxQuantityOrdered") not in gap_keys(gaps)

    def test_float_exported_value_equals_integer_ideal(self):
        """Oracle exports tolerances as floats (5.0); ideal stored as '5' → should NOT gap."""
        cip, gaps, _, _ = run(
            BUS_INDIA,
            {
                "MaxQuantityOrdered": ["5.0"] * 5,  # Oracle float export
                "MaxQuantityReceived": ["5.0"] * 5,
                "MaxAmountOrdered": ["10.0"] * 5,
                "MaxAmountReceived": ["10.0"] * 5,
                "TaxAmountRange": ["500"] * 5,
            },
            TOL_IDEAL,
        )
        assert len(gaps) == 0, (
            f"Float normalization failed: {[(g.bu_name, g.config_name, g.actual_value) for g in gaps]}"
        )

    def test_zero_tolerance_is_gap(self):
        """Tolerance set to 0 is below the expected 5 — control gap."""
        cip, gaps, _, _ = run(
            ["Mumbai Plant"],
            {
                "MaxQuantityOrdered": ["0"],
                "MaxQuantityReceived": ["5"],
                "MaxAmountOrdered": ["10"],
                "MaxAmountReceived": ["10"],
                "TaxAmountRange": ["500"],
            },
            TOL_IDEAL,
        )
        assert ("Mumbai Plant", "MaxQuantityOrdered") in gap_keys(gaps)

    def test_tax_amount_range_is_captured(self):
        """TaxAmountRange is capture — must not appear in gaps regardless of value."""
        cip, gaps, extra, _ = run(
            BUS_INDIA,
            {
                "MaxQuantityOrdered": ["5"] * 5,
                "MaxQuantityReceived": ["5"] * 5,
                "MaxAmountOrdered": ["10"] * 5,
                "MaxAmountReceived": ["10"] * 5,
                "TaxAmountRange": ["100", "200", "300", "400", "500"],
            },
            TOL_IDEAL,
        )
        assert all(r.config_name != "TaxAmountRange" for r in gaps)
        tax_extra = [r for r in extra if r.config_name == "TaxAmountRange"]
        assert len(tax_extra) == len(BUS_INDIA)
        assert {r.actual_value for r in tax_extra} == {
            "100",
            "200",
            "300",
            "400",
            "500",
        }

    def test_multiple_tolerance_violations_across_bus(self):
        """Each BU with ANY tolerance above limit must be individually flagged."""
        cip, gaps, _, _ = run(
            BUS_INDIA,
            {
                "MaxQuantityOrdered": ["5", "10", "15", "5", "20"],
                "MaxQuantityReceived": ["5"] * 5,
                "MaxAmountOrdered": ["10"] * 5,
                "MaxAmountReceived": ["10"] * 5,
                "TaxAmountRange": ["500"] * 5,
            },
            TOL_IDEAL,
        )
        violating = {"Delhi HQ", "Bangalore IT", "Pune SEZ"}
        compliant = {"Mumbai Plant", "Chennai Plant"}
        for bu in violating:
            assert (bu, "MaxQuantityOrdered") in gap_keys(gaps), f"{bu} should be a gap"
        for bu in compliant:
            assert (bu, "MaxQuantityOrdered") not in gap_keys(gaps), (
                f"{bu} should be CIP"
            )


# ══════════════════════════════════════════════════════════════════════════════
# MODULE 5 — AP Common Pay Proc Options
# ══════════════════════════════════════════════════════════════════════════════


PAY_IDEAL = [
    ("PositivePayRequired", "Y", "Y, N"),
    ("AutoCalculateTaxFlag", "Y", "Y, N"),
    ("PrepaymentPayGroup", "PREPAY", ""),
    ("DefaultExchangeRateType", "Corporate", "Corporate, Spot, User"),
    ("PaymentTerms", "Date", ""),  # capture — type descriptor
    ("BankChargeBearer", "Date", ""),  # capture
]


class TestAPCommonPayProcOptions:
    def test_positive_pay_disabled_is_gap(self):
        """PositivePayRequired=N means bank fraud protection is off — critical gap."""
        cip, gaps, _, _ = run(
            BUS_INDIA,
            {
                "PositivePayRequired": ["Y", "N", "Y", "Y", "Y"],  # Delhi HQ off ← gap
                "AutoCalculateTaxFlag": ["Y"] * 5,
                "PrepaymentPayGroup": ["PREPAY"] * 5,
                "DefaultExchangeRateType": ["Corporate"] * 5,
                "PaymentTerms": ["Net30"] * 5,
                "BankChargeBearer": ["BUYER"] * 5,
            },
            PAY_IDEAL,
        )
        assert ("Delhi HQ", "PositivePayRequired") in gap_keys(gaps)
        assert ("Mumbai Plant", "PositivePayRequired") not in gap_keys(gaps)

    def test_spot_rate_type_is_gap(self):
        """DefaultExchangeRateType=Spot (manual) instead of Corporate — gap."""
        cip, gaps, _, _ = run(
            ["Mumbai Plant", "Delhi HQ"],
            {
                "PositivePayRequired": ["Y", "Y"],
                "AutoCalculateTaxFlag": ["Y", "Y"],
                "PrepaymentPayGroup": ["PREPAY", "PREPAY"],
                "DefaultExchangeRateType": [
                    "Corporate",
                    "Spot",
                ],  # Delhi HQ = Spot ← gap
                "PaymentTerms": ["Net30", "Net30"],
                "BankChargeBearer": ["BUYER", "BUYER"],
            },
            PAY_IDEAL,
        )
        assert ("Delhi HQ", "DefaultExchangeRateType") in gap_keys(gaps)
        assert ("Mumbai Plant", "DefaultExchangeRateType") not in gap_keys(gaps)

    def test_wrong_prepayment_group_is_gap(self):
        cip, gaps, _, _ = run(
            ["Mumbai Plant"],
            {
                "PositivePayRequired": ["Y"],
                "AutoCalculateTaxFlag": ["Y"],
                "PrepaymentPayGroup": ["STANDARD"],  # should be PREPAY ← gap
                "DefaultExchangeRateType": ["Corporate"],
                "PaymentTerms": ["Net30"],
                "BankChargeBearer": ["BUYER"],
            },
            PAY_IDEAL,
        )
        assert ("Mumbai Plant", "PrepaymentPayGroup") in gap_keys(gaps)

    def test_payment_terms_and_bank_charge_are_captured(self):
        """Both 'Date' type-descriptor fields are captured, never compared."""
        cip, gaps, extra, _ = run(
            BUS_INDIA,
            {
                "PositivePayRequired": ["Y"] * 5,
                "AutoCalculateTaxFlag": ["Y"] * 5,
                "PrepaymentPayGroup": ["PREPAY"] * 5,
                "DefaultExchangeRateType": ["Corporate"] * 5,
                "PaymentTerms": ["Net30", "Net45", "Net60", "Net30", "Net90"],
                "BankChargeBearer": ["BUYER"] * 5,
            },
            PAY_IDEAL,
        )
        assert all(
            r.config_name not in {"PaymentTerms", "BankChargeBearer"} for r in gaps
        )
        captured_names = {r.config_name for r in extra}
        assert "PaymentTerms" in captured_names
        assert "BankChargeBearer" in captured_names

    def test_all_pay_proc_compliant_zero_gaps(self):
        cip, gaps, _, _ = run(
            BUS_INDIA,
            {
                "PositivePayRequired": ["Y"] * 5,
                "AutoCalculateTaxFlag": ["Y"] * 5,
                "PrepaymentPayGroup": ["PREPAY"] * 5,
                "DefaultExchangeRateType": ["Corporate"] * 5,
                "PaymentTerms": ["Net30"] * 5,
                "BankChargeBearer": ["BUYER"] * 5,
            },
            PAY_IDEAL,
        )
        assert len(gaps) == 0


# ══════════════════════════════════════════════════════════════════════════════
# Cross-cutting audit rules
# ══════════════════════════════════════════════════════════════════════════════


class TestCrossCuttingAuditRules:
    def test_every_bu_gets_its_own_row_in_report(self):
        """Each BU × config combination is a separate audit finding — no aggregation."""
        n = 10
        bus = [f"Plant_{i:02d}" for i in range(n)]
        cip, gaps, _, _ = run(
            bus,
            {"InventoryFlag": ["Y"] * 8 + ["N", "N"]},
            [("InventoryFlag", "Y", "Y, N")],
        )
        assert len(cip) == 8
        assert len(gaps) == 2
        assert {r.bu_name for r in gaps} == {"Plant_08", "Plant_09"}

    def test_unmatched_config_never_produces_findings(self):
        """A config with no matching column in the actual file must produce zero rows."""
        cip, gaps, extra, mapping = run(
            BUS_INDIA,
            {"InventoryFlag": ["Y"] * 5},
            [("LastUpdateDate", "Date", ""), ("InventoryFlag", "Y", "Y, N")],
        )
        lastupdate_findings = [
            r for r in cip + gaps + extra if r.config_name == "LastUpdateDate"
        ]
        assert len(lastupdate_findings) == 0

    def test_comment_text_matches_sheet_name(self):
        """Row comments must exactly match the sheet they land in."""
        cip, gaps, extra, _ = run(
            ["Mumbai Plant"],
            {
                "InventoryFlag": ["Y"],
                "AllowNegOnhandCcTxns": ["Y"],
                "OrganizationCode": [1],
            },
            [
                ("InventoryFlag", "Y", ""),
                ("AllowNegOnhandCcTxns", "N", ""),
                ("OrganizationCode", "Capture", ""),
            ],
        )
        assert cip[0].comment == "Controls in place"
        assert gaps[0].comment == "Controls gaps"
        assert extra[0].comment == "Actual config captured"

    def test_options_string_in_gap_matches_ideal_file(self):
        """The options string on a gap row must come from the ideal file, not be empty."""
        cip, gaps, extra, _ = run(
            ["Delhi HQ"],
            {"InvoiceMatchOption": ["Receipt"]},
            [("InvoiceMatchOption", "PO", "PO, Receipt, Order")],
        )
        assert len(gaps) == 1
        assert gaps[0].options == "PO, Receipt, Order"

    def test_report_has_five_sheets(self):
        """The Excel report always has exactly 5 sheets regardless of module."""
        import io as _io
        import openpyxl

        actual_df = pd.DataFrame(
            {
                "BU_NAME": ["Mumbai Plant"],
                "InventoryFlag": ["Y"],
            }
        )
        ideal_df = pd.DataFrame(
            {
                "config_name": ["InventoryFlag"],
                "ideal_value": ["Y"],
                "options": ["Y, N"],
            }
        )
        mapping = build_mapping(ideal_df, list(actual_df.columns))
        cip, gaps, extra = compare(actual_df, mapping)
        xlsx = build_report(
            cip,
            gaps,
            extra,
            mapping,
            actual_filename="client.xlsx",
            ideal_filename="ideal.xlsx",
            total_bu_rows=1,
            fuzzy_threshold=80,
        )
        wb = openpyxl.load_workbook(_io.BytesIO(xlsx))
        assert len(wb.sheetnames) == 5
        assert set(wb.sheetnames) == {
            "Controls in place",
            "Control gaps",
            "Controls additional data",
            "Header Mapping and Exc",
            "Audit Summary",
        }

    def test_control_gaps_sheet_has_valid_options_column(self):
        """Control gaps sheet must include the Valid Options column for auditor context."""
        import io as _io
        import openpyxl

        actual_df = pd.DataFrame(
            {
                "BU_NAME": ["Delhi HQ"],
                "InvoiceMatchOption": ["Receipt"],
            }
        )
        ideal_df = pd.DataFrame(
            {
                "config_name": ["InvoiceMatchOption"],
                "ideal_value": ["PO"],
                "options": ["PO, Receipt, Order"],
            }
        )
        mapping = build_mapping(ideal_df, list(actual_df.columns))
        cip, gaps, extra = compare(actual_df, mapping)
        xlsx = build_report(
            cip,
            gaps,
            extra,
            mapping,
            actual_filename="client.xlsx",
            ideal_filename="ideal.xlsx",
            total_bu_rows=1,
            fuzzy_threshold=80,
        )
        wb = openpyxl.load_workbook(_io.BytesIO(xlsx))
        ws = wb["Control gaps"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "Valid Options" in headers
        assert "Ideal Value" in headers


# ══════════════════════════════════════════════════════════════════════════════
# TC-1  Integer LOV codes
# Known limitation: normalize_value("1") → "y" via bool map. LOV codes that
# happen to use "1" or "0" as values are indistinguishable from booleans.
# ══════════════════════════════════════════════════════════════════════════════


class TestIntegerLOVCodes:
    """Document the known bool-map collision on integer LOV codes."""

    def test_lov_code_1_matches_ideal_1_as_cip(self):
        """ReceivingRoutingId ideal='1', actual='1' → CIP (correct result)."""
        cip, gaps, _, _ = run(
            ["Mumbai Plant"],
            {"ReceivingRoutingId": ["1"]},
            [("ReceivingRoutingId", "1", "1-Direct, 2-Standard, 3-Inspection")],
        )
        assert ("Mumbai Plant", "ReceivingRoutingId") in cip_keys(cip)

    def test_lov_code_2_vs_ideal_1_is_gap(self):
        """ReceivingRoutingId actual='2' (Standard) when ideal='1' (Direct) → gap."""
        cip, gaps, _, _ = run(
            ["Delhi HQ"],
            {"ReceivingRoutingId": ["2"]},
            [("ReceivingRoutingId", "1", "1-Direct, 2-Standard, 3-Inspection")],
        )
        assert ("Delhi HQ", "ReceivingRoutingId") in gap_keys(gaps)

    def test_boolean_true_collides_with_lov_code_1(self):
        # Known bug: normalize("true") == "y" == normalize("1").
        # A boolean "true" export value is treated as equivalent to LOV code "1".
        # This is a false CIP — the actual value is semantically different.
        from engine.normalizer import normalize_value

        assert normalize_value("true") == normalize_value("1") == "y"

        cip, gaps, _, _ = run(
            ["Mumbai Plant"],
            {"ReceivingRoutingId": ["true"]},  # mis-exported as boolean
            [("ReceivingRoutingId", "1", "1-Direct, 2-Standard, 3-Inspection")],
        )
        # Current behaviour: false CIP (both map to "y"). Asserted here to
        # track the known limitation — fix via per-config comparison mode.
        assert ("Mumbai Plant", "ReceivingRoutingId") in cip_keys(cip)

    def test_lov_code_0_maps_to_n_not_zero(self):
        """LOV code '0' normalizes to 'n'. ideal='0', actual='0' → CIP via 'n'=='n'."""
        from engine.normalizer import normalize_value

        assert normalize_value("0") == "n"

        cip, gaps, _, _ = run(
            ["Mumbai Plant"],
            {"NegativeInvReceiptCode": ["0"]},
            [("NegativeInvReceiptCode", "0", "0-Allowed, 1-Not Allowed")],
        )
        assert ("Mumbai Plant", "NegativeInvReceiptCode") in cip_keys(cip)


# ══════════════════════════════════════════════════════════════════════════════
# TC-2  Unicode BU names
# Arabic, CJK, and accented characters must survive the stringify/strip pipeline.
# ══════════════════════════════════════════════════════════════════════════════


class TestUnicodeBUNames:
    UNICODE_BUS = [
        "مصنع دبي",  # Arabic  — Dubai Factory
        "東京工場",  # Japanese — Tokyo Factory
        "上海制造",  # Chinese  — Shanghai Manufacturing
        "Münih Tesisi",  # Turkish  — Munich Facility (accented)
        "Nairobi Plant",  # ASCII control
    ]

    def test_unicode_bu_names_appear_in_cip(self):
        """All BU names (inc. Arabic/CJK) must be preserved exactly in CIP rows."""
        cip, gaps, _, _ = run(
            self.UNICODE_BUS,
            {"InventoryFlag": ["Y"] * 5},
            [("InventoryFlag", "Y", "Y, N")],
        )
        cip_bus = {r.bu_name for r in cip}
        for bu in self.UNICODE_BUS:
            assert bu in cip_bus, f"BU name lost: {bu!r}"

    def test_unicode_bu_names_appear_in_gaps(self):
        """Control gaps must carry the original Unicode BU name unmodified."""
        cip, gaps, _, _ = run(
            self.UNICODE_BUS,
            {"InventoryFlag": ["N"] * 5},
            [("InventoryFlag", "Y", "Y, N")],
        )
        gap_bus = {r.bu_name for r in gaps}
        for bu in self.UNICODE_BUS:
            assert bu in gap_bus, f"BU name lost in gaps: {bu!r}"

    def test_mixed_unicode_and_ascii_bus_produce_correct_row_count(self):
        """Each BU gets exactly one row per config — no merging or skipping."""
        cip, gaps, _, _ = run(
            self.UNICODE_BUS,
            {"InventoryFlag": ["Y", "N", "Y", "N", "Y"]},
            [("InventoryFlag", "Y", "Y, N")],
        )
        assert len(cip) == 3
        assert len(gaps) == 2


# ══════════════════════════════════════════════════════════════════════════════
# TC-3  Duplicate BU rows
# When a BU appears more than once in the export (split organizations, test
# data), each row is an independent finding — never merged or deduplicated.
# ══════════════════════════════════════════════════════════════════════════════


class TestDuplicateBURows:
    def test_same_bu_twice_both_compliant_produces_two_cip_rows(self):
        """Two identical BU rows, both CIP → 2 CIP rows, 0 gaps."""
        cip, gaps, _, _ = run(
            ["Mumbai Plant", "Mumbai Plant"],
            {"InventoryFlag": ["Y", "Y"]},
            [("InventoryFlag", "Y", "Y, N")],
        )
        assert len(cip) == 2
        assert len(gaps) == 0

    def test_same_bu_twice_one_gap_one_cip(self):
        """Two rows for same BU with different values → one CIP and one gap."""
        cip, gaps, _, _ = run(
            ["Mumbai Plant", "Mumbai Plant"],
            {"InventoryFlag": ["Y", "N"]},
            [("InventoryFlag", "Y", "Y, N")],
        )
        assert len(cip) == 1
        assert len(gaps) == 1
        assert cip[0].bu_name == "Mumbai Plant"
        assert gaps[0].bu_name == "Mumbai Plant"

    def test_triplicate_bu_all_gaps(self):
        """Three rows for the same BU, all non-compliant → 3 separate gap rows."""
        cip, gaps, _, _ = run(
            ["Delhi HQ", "Delhi HQ", "Delhi HQ"],
            {"InventoryFlag": ["N", "N", "N"]},
            [("InventoryFlag", "Y", "Y, N")],
        )
        assert len(gaps) == 3
        assert all(r.bu_name == "Delhi HQ" for r in gaps)

    def test_duplicate_capture_rows_both_extracted(self):
        """Duplicate BU rows for a capture config → both values extracted independently."""
        cip, gaps, extra, _ = run(
            ["Mumbai Plant", "Mumbai Plant"],
            {"OrganizationCode": [2, 999]},
            [("OrganizationCode", "Capture", "")],
        )
        assert len(extra) == 2
        values = [r.actual_value for r in extra]
        assert "2" in values
        assert "999" in values


# ══════════════════════════════════════════════════════════════════════════════
# TC-4  Fuzzy false-positive boundary
# Matches in the 80–89 range must carry the "Low confidence — verify mapping"
# remark so auditors know to manually confirm the column alignment.
# ══════════════════════════════════════════════════════════════════════════════


class TestFuzzyFalsePositiveBoundary:
    def test_score_below_90_gets_low_confidence_remark(self):
        """Fuzzy score 80–89 → MappingResult.remarks contains 'Low confidence'."""
        from thefuzz import fuzz

        ideal_name = "ReceivingRoutingFlag"
        actual_name = "ReceivingRoutingId"
        score = fuzz.token_sort_ratio(ideal_name.lower(), actual_name.lower())
        assert 80 <= score < 90, (
            f"Precondition failed: score={score}. Pick a different pair if fuzzy lib changed."
        )

        ideal_df = pd.DataFrame(
            {"config_name": [ideal_name], "ideal_value": ["1"], "options": [""]}
        )
        mapping = build_mapping(ideal_df, ["BU_NAME", actual_name], fuzzy_threshold=80)
        m = mapping[0]
        assert m.status == "Matched"
        assert m.match_method == "Fuzzy"
        assert "Low confidence" in m.remarks

    def test_score_at_or_above_90_has_no_remark(self):
        """Fuzzy score >= 90 → MappingResult.remarks is empty (high confidence)."""
        from thefuzz import fuzz

        ideal_name = "AllowItemSubstitutionFlag"
        actual_name = "AllowItemSubstitutionsFlag"
        score = fuzz.token_sort_ratio(ideal_name.lower(), actual_name.lower())
        assert score >= 90, (
            f"Precondition failed: score={score}. Pick a closer pair if fuzzy lib changed."
        )

        ideal_df = pd.DataFrame(
            {"config_name": [ideal_name], "ideal_value": ["Y"], "options": [""]}
        )
        mapping = build_mapping(ideal_df, ["BU_NAME", actual_name], fuzzy_threshold=80)
        m = mapping[0]
        assert m.status == "Matched"
        assert m.match_method == "Fuzzy"
        assert m.remarks == ""

    def test_below_threshold_is_unmatched_not_low_confidence(self):
        """Score below threshold → Unmatched, not a low-confidence match."""
        ideal_df = pd.DataFrame(
            {
                "config_name": ["LastUpdateDate"],
                "ideal_value": ["Date"],
                "options": [""],
            }
        )
        mapping = build_mapping(
            ideal_df,
            ["BU_NAME", "InventoryFlag", "ReceivingRoutingId"],
            fuzzy_threshold=80,
        )
        m = mapping[0]
        assert m.status == "Unmatched"
        assert "Low confidence" not in m.remarks


# ══════════════════════════════════════════════════════════════════════════════
# TC-5  All-null actual column
# When every BU has a null/NaN value for a comparison config, the engine must
# produce one control gap per BU — never a false CIP.
# ══════════════════════════════════════════════════════════════════════════════


class TestAllNullActualColumn:
    def test_all_null_column_produces_all_gaps(self):
        """10 BUs, all None for InventoryFlag → 10 gaps, 0 CIPs."""
        bus = [f"Plant_{i:02d}" for i in range(10)]
        cip, gaps, extra, _ = run(
            bus,
            {"InventoryFlag": [None] * 10},
            [("InventoryFlag", "Y", "Y, N")],
        )
        assert len(cip) == 0
        assert len(gaps) == 10
        assert {r.bu_name for r in gaps} == set(bus)

    def test_all_nan_column_produces_all_gaps(self):
        """float('nan') values (pandas default for missing cells) → all gaps."""

        bus = ["Mumbai Plant", "Delhi HQ", "Bangalore IT"]
        cip, gaps, _, _ = run(
            bus,
            {"InventoryFlag": [float("nan")] * 3},
            [("InventoryFlag", "Y", "Y, N")],
        )
        assert len(gaps) == 3
        assert len(cip) == 0
        assert all(r.actual_value == "" for r in gaps)

    def test_mixed_null_and_valid_values(self):
        """Some nulls → gaps; non-null matching values → CIP. Counts must add up."""
        bus = ["Plant_A", "Plant_B", "Plant_C", "Plant_D"]
        cip, gaps, _, _ = run(
            bus,
            {"InventoryFlag": ["Y", None, "Y", None]},
            [("InventoryFlag", "Y", "Y, N")],
        )
        assert len(cip) == 2
        assert len(gaps) == 2
        assert {r.bu_name for r in gaps} == {"Plant_B", "Plant_D"}

    def test_all_null_capture_column_produces_empty_actual_values(self):
        """All-null capture column → extra rows with empty actual_value strings."""
        bus = ["Mumbai Plant", "Delhi HQ"]
        cip, gaps, extra, _ = run(
            bus,
            {"OrganizationCode": [None, None]},
            [("OrganizationCode", "Capture", "")],
        )
        assert len(extra) == 2
        assert all(r.actual_value == "" for r in extra)
        assert len(gaps) == 0
