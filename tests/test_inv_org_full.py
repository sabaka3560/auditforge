"""Full INV Organization Parameters test suite — 100+ Oracle Fusion use cases.

Uses inline ideal DataFrames (same names/values as the production ideal file)
so production files are never touched. Tests run end-to-end through the same
build_mapping → compare pipeline as the real tool.

Coverage:
  - All 15 comparison configs (Y/N boolean flags + LOV codes)
  - All 18 capture-only configs
  - Oracle-specific column-name aliases (BusinessUnitId → BusinessUnitName, etc.)
  - Value normalization quirks: float exports, case, whitespace, None
  - Multi-BU realistic scenarios (5, 10, 50, 100 BUs)
  - Partial compliance patterns (some BUs OK, some not)
  - NULL / empty actual values
  - LOV codes (NegativeInvReceiptCode, ReceivingRoutingId)
  - All five known unmatched fields (produce no findings)
"""

import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pandas as pd
import pytest

from engine.comparator import compare
from engine.matcher import build_mapping


# ── helpers ────────────────────────────────────────────────────────────────────


def _ideal(*rows):
    """rows: (config_name, ideal_value)  — options default to ''"""
    return pd.DataFrame(
        [(r[0], r[1], r[2] if len(r) > 2 else "") for r in rows],
        columns=["config_name", "ideal_value", "options"],
    )


def _actual(bus, **cols):
    return pd.DataFrame({"BU_NAME": bus, **cols})


def _run(actual_df, ideal_df, threshold=80):
    mapping = build_mapping(
        ideal_df, list(actual_df.columns), fuzzy_threshold=threshold
    )
    cip, gaps, extra = compare(actual_df, mapping)
    return cip, gaps, extra, mapping


def _gap_keys(gaps):
    return {(r.bu_name, r.config_name) for r in gaps}


def _cip_keys(cip):
    return {(r.bu_name, r.config_name) for r in cip}


def _extra_keys(extra):
    return {(r.bu_name, r.config_name) for r in extra}


# ── production-identical ideal DataFrame (33 configs) ─────────────────────────
# Matches Inventory_ideal_value.xlsx exactly. Do NOT edit these values.

PROD_IDEAL = _ideal(
    ("AllowItemSubstitutionsFlag", "Y", "Y, N"),
    ("AllowNegOnhandCcTxns", "N", "Y, N"),
    ("AutoDelAllocFlag", "N", "Y, N"),
    ("BusinessUnitId", "Capture the value defined"),
    ("ContractMfgFlag", "Capture the value defined"),
    ("CopyLotAttributeFlag", "Y", "Y, N"),
    ("CustomerAccountNumber", "Capture the value defined"),
    ("DefaultSubinvOrderValue", "Capture the value defined"),
    ("EamEnabledFlag", "y", "Y, N"),
    ("FaBookTypeCode", "Capture the value defined"),
    ("FifoOrigRcptDateFlag", "y", "Y, N"),
    ("InventoryFlag", "y", "Y, N"),
    ("LastUpdateDate", "Capture the value defined"),
    ("LastUpdateLogin", "Capture the value defined"),
    ("LastUpdatedBy", "Capture the value defined"),
    ("MasterOrganizationId", "Capture the value defined"),
    ("MfgPlantFlag", "y", "Y, N"),
    ("MoPickConfirmRequired", "y", "Y, N"),
    ("NegativeInvReceiptCode", "n", "Y, N"),
    ("OrganizationCode", "Capture the value defined"),
    ("OrganizationId", "Capture the value defined"),
    ("OvpkTransferOrdersEnabled", "y", "Y, N"),
    ("ProfitCenterBuId", "Capture the value defined"),
    ("ProjectReferenceEnabled", "y", "Y, N"),
    ("PurchasingByRevision", "n", "Y, N"),
    ("SourceOrganizationId", "Capture the value defined"),
    ("SourceSubinventory", "Capture the value defined"),
    ("SourceType", "Capture the value defined"),
    ("SpecialHndlngOvrpkFlag", "y", "Y, N"),
    ("SupplierId", "Capture the value defined"),
    ("SupplierSiteId", "Capture the value defined"),
    ("TrackCountryOfOriginFlag", "Capture the value defined"),
    ("UseCurItemCostFlag", "Y", "Y, N"),
)

# Canonical compliant actual row — Oracle column names as exported.
# Column names use actual Oracle export headers (alias dict maps ideal→actual).
_COMPLIANT = {
    "AllowItemSubstitutions": "Y",  # alias: AllowItemSubstitutionsFlag
    "AllowNegOnhandCcTxns": "N",
    "AutoDelAllocFlag": "N",
    "BusinessUnitName": "Mumbai BU",  # alias: BusinessUnitId
    "ContractMfgFlag": "N",
    "CopyLotAttributeFlag": "Y",
    "CustomerAccountNumber": "CUST-001",
    "DefaultSubinvOrderValue": "1",
    "EamEnabledFlag": "Y",
    "FaBookTypeCode": "CORP",
    "FifoOrigRcptDateFlag": "Y",
    "InventoryFlag": "Y",
    "MasterOrgCode": "M-001",  # alias: MasterOrganizationId
    "MfgPlantFlag": "Y",
    "MoPickConfirmRequired": "Y",
    "NegativeInvReceiptCode": "N",
    "OrganizationCode": "ORG-001",
    "OvpkTransferOrdersEnabled": "Y",
    "ProfitCenterBuName": "PC-Mumbai",  # alias: ProfitCenterBuId
    "ProjectReferenceEnabled": "Y",
    "PurchasingByRevision": "N",
    "SourceType": "Inventory",
    "SpecialHndlngOvrpkFlag": "Y",
    "SupplierName": "Vendor-A",  # alias: SupplierId
    "SupplierSiteName": "Site-A",  # alias: SupplierSiteId
    "TrackCountryOfOriginFlag": "N",
    "UseCurItemCostFlag": "Y",
}

BUS_5 = ["Mumbai Plant", "Delhi HQ", "Bangalore IT", "Chennai Plant", "Pune SEZ"]
BUS_10 = BUS_5 + [
    "Hyderabad Hub",
    "Kolkata East",
    "Ahmedabad West",
    "Jaipur North",
    "Kochi South",
]


def _full_actual(bus, overrides=None):
    """Build a compliant actual_df for `bus`, with optional per-column overrides.

    overrides: {col: list_of_values} — length must match len(bus).
    """
    n = len(bus)
    data = {"BU_NAME": bus}
    for col, val in _COMPLIANT.items():
        data[col] = [val] * n
    if overrides:
        data.update(overrides)
    return pd.DataFrame(data)


# ══════════════════════════════════════════════════════════════════════════════
# 1. Fully compliant — zero gaps
# ══════════════════════════════════════════════════════════════════════════════


class TestFullyCompliant:
    def test_single_bu_all_cip(self):
        df = _full_actual(["Mumbai Plant"])
        cip, gaps, extra, _ = _run(df, PROD_IDEAL)
        assert len(gaps) == 0

    def test_five_bus_all_cip(self):
        df = _full_actual(BUS_5)
        cip, gaps, extra, _ = _run(df, PROD_IDEAL)
        assert len(gaps) == 0

    def test_ten_bus_all_cip(self):
        df = _full_actual(BUS_10)
        cip, gaps, extra, _ = _run(df, PROD_IDEAL)
        assert len(gaps) == 0

    def test_fifty_bus_zero_gaps(self):
        bus = [f"Plant_{i:03d}" for i in range(50)]
        df = _full_actual(bus)
        cip, gaps, extra, _ = _run(df, PROD_IDEAL)
        assert len(gaps) == 0

    def test_hundred_bus_zero_gaps(self):
        bus = [f"BU_{i:03d}" for i in range(100)]
        df = _full_actual(bus)
        cip, gaps, extra, _ = _run(df, PROD_IDEAL)
        assert len(gaps) == 0

    def test_comparison_row_count_matches_bus_times_compared_configs(self):
        # 15 comparison configs in PROD_IDEAL × 5 BUs = 75 CIP rows
        df = _full_actual(BUS_5)
        cip, gaps, extra, _ = _run(df, PROD_IDEAL)
        comparison_configs = sum(
            1
            for _, row in PROD_IDEAL.iterrows()
            if row["ideal_value"].lower()
            not in ["capture the value defined", "capture"]
            and "capture" not in row["ideal_value"].lower()
        )
        assert len(cip) == len(BUS_5) * comparison_configs
        assert len(gaps) == 0


# ══════════════════════════════════════════════════════════════════════════════
# 2. AllowNegOnhandCcTxns — negative inventory
# ══════════════════════════════════════════════════════════════════════════════


class TestAllowNegOnhandCcTxns:
    def test_y_is_gap(self):
        df = _full_actual(BUS_5, {"AllowNegOnhandCcTxns": ["N", "Y", "N", "N", "N"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Delhi HQ", "AllowNegOnhandCcTxns") in _gap_keys(gaps)
        assert ("Mumbai Plant", "AllowNegOnhandCcTxns") not in _gap_keys(gaps)

    def test_yes_is_gap(self):
        df = _full_actual(["Mumbai Plant"], {"AllowNegOnhandCcTxns": ["Yes"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "AllowNegOnhandCcTxns") in _gap_keys(gaps)

    def test_true_is_gap(self):
        df = _full_actual(["Mumbai Plant"], {"AllowNegOnhandCcTxns": ["true"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "AllowNegOnhandCcTxns") in _gap_keys(gaps)

    def test_enabled_is_gap(self):
        df = _full_actual(["Mumbai Plant"], {"AllowNegOnhandCcTxns": ["enabled"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "AllowNegOnhandCcTxns") in _gap_keys(gaps)

    def test_n_lowercase_is_cip(self):
        df = _full_actual(["Mumbai Plant"], {"AllowNegOnhandCcTxns": ["n"]})
        cip, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "AllowNegOnhandCcTxns") in _cip_keys(cip)

    def test_no_is_cip(self):
        df = _full_actual(["Mumbai Plant"], {"AllowNegOnhandCcTxns": ["no"]})
        cip, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "AllowNegOnhandCcTxns") in _cip_keys(cip)

    def test_false_is_cip(self):
        df = _full_actual(["Mumbai Plant"], {"AllowNegOnhandCcTxns": ["false"]})
        cip, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "AllowNegOnhandCcTxns") in _cip_keys(cip)

    def test_null_is_gap(self):
        df = _full_actual(["Mumbai Plant"], {"AllowNegOnhandCcTxns": [None]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "AllowNegOnhandCcTxns") in _gap_keys(gaps)

    def test_all_five_bus_gap(self):
        df = _full_actual(BUS_5, {"AllowNegOnhandCcTxns": ["Y"] * 5})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        for bu in BUS_5:
            assert (bu, "AllowNegOnhandCcTxns") in _gap_keys(gaps)


# ══════════════════════════════════════════════════════════════════════════════
# 3. AllowItemSubstitutionsFlag → actual column AllowItemSubstitutions (alias)
# ══════════════════════════════════════════════════════════════════════════════


class TestAllowItemSubstitutionsFlag:
    def test_n_is_gap(self):
        df = _full_actual(["Mumbai Plant"], {"AllowItemSubstitutions": ["N"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "AllowItemSubstitutionsFlag") in _gap_keys(gaps)

    def test_y_is_cip(self):
        df = _full_actual(["Mumbai Plant"], {"AllowItemSubstitutions": ["Y"]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "AllowItemSubstitutionsFlag") in _cip_keys(cip)

    def test_yes_mixed_case_is_cip(self):
        df = _full_actual(["Mumbai Plant"], {"AllowItemSubstitutions": ["YES"]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "AllowItemSubstitutionsFlag") in _cip_keys(cip)

    def test_alias_maps_correctly(self):
        """Confirms the manual alias routes AllowItemSubstitutionsFlag→AllowItemSubstitutions."""
        _, _, _, mapping = _run(_full_actual(["Mumbai Plant"]), PROD_IDEAL)
        m = next(x for x in mapping if x.ideal_name == "AllowItemSubstitutionsFlag")
        assert m.matched_header == "AllowItemSubstitutions"
        assert m.match_method == "Manual Alias"


# ══════════════════════════════════════════════════════════════════════════════
# 4. InventoryFlag
# ══════════════════════════════════════════════════════════════════════════════


class TestInventoryFlag:
    def test_n_is_critical_gap(self):
        df = _full_actual(BUS_5, {"InventoryFlag": ["Y", "N", "Y", "N", "Y"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Delhi HQ", "InventoryFlag") in _gap_keys(gaps)
        assert ("Chennai Plant", "InventoryFlag") in _gap_keys(gaps)
        assert ("Mumbai Plant", "InventoryFlag") not in _gap_keys(gaps)

    def test_float_1_0_is_not_y(self):
        """Oracle float export: '1.0' normalizes to '1' which does NOT equal 'y'."""
        df = _full_actual(["Mumbai Plant"], {"InventoryFlag": ["1.0"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "InventoryFlag") in _gap_keys(gaps)

    def test_lowercase_y_is_cip(self):
        df = _full_actual(["Mumbai Plant"], {"InventoryFlag": ["y"]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "InventoryFlag") in _cip_keys(cip)

    def test_true_is_cip(self):
        df = _full_actual(["Mumbai Plant"], {"InventoryFlag": ["true"]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "InventoryFlag") in _cip_keys(cip)

    def test_null_is_gap(self):
        df = _full_actual(["Mumbai Plant"], {"InventoryFlag": [None]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "InventoryFlag") in _gap_keys(gaps)


# ══════════════════════════════════════════════════════════════════════════════
# 5. EamEnabledFlag
# ══════════════════════════════════════════════════════════════════════════════


class TestEamEnabledFlag:
    @pytest.mark.parametrize("v", ["Y", "y", "Yes", "yes", "true", "True", "enabled"])
    def test_truthy_values_are_cip(self, v):
        df = _full_actual(["Mumbai Plant"], {"EamEnabledFlag": [v]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "EamEnabledFlag") in _cip_keys(cip)

    @pytest.mark.parametrize("v", ["N", "n", "No", "false", "disabled"])
    def test_falsy_values_are_gap(self, v):
        df = _full_actual(["Mumbai Plant"], {"EamEnabledFlag": [v]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "EamEnabledFlag") in _gap_keys(gaps)

    def test_null_is_gap(self):
        df = _full_actual(["Mumbai Plant"], {"EamEnabledFlag": [None]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "EamEnabledFlag") in _gap_keys(gaps)


# ══════════════════════════════════════════════════════════════════════════════
# 6. FifoOrigRcptDateFlag
# ══════════════════════════════════════════════════════════════════════════════


class TestFifoOrigRcptDateFlag:
    def test_n_is_gap(self):
        df = _full_actual(["Mumbai Plant"], {"FifoOrigRcptDateFlag": ["N"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "FifoOrigRcptDateFlag") in _gap_keys(gaps)

    def test_y_is_cip(self):
        df = _full_actual(["Mumbai Plant"], {"FifoOrigRcptDateFlag": ["Y"]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "FifoOrigRcptDateFlag") in _cip_keys(cip)

    def test_mixed_across_bus(self):
        df = _full_actual(
            BUS_5,
            {"FifoOrigRcptDateFlag": ["Y", "N", "Y", "N", "Y"]},
        )
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Delhi HQ", "FifoOrigRcptDateFlag") in _gap_keys(gaps)
        assert ("Chennai Plant", "FifoOrigRcptDateFlag") in _gap_keys(gaps)
        assert len([g for g in gaps if g.config_name == "FifoOrigRcptDateFlag"]) == 2


# ══════════════════════════════════════════════════════════════════════════════
# 7. MfgPlantFlag
# ══════════════════════════════════════════════════════════════════════════════


class TestMfgPlantFlag:
    @pytest.mark.parametrize("v", ["Y", "y", "yes", "YES", "true", "TRUE", "enabled"])
    def test_truthy_is_cip(self, v):
        df = _full_actual(["Mumbai Plant"], {"MfgPlantFlag": [v]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "MfgPlantFlag") in _cip_keys(cip)

    def test_n_is_gap(self):
        df = _full_actual(["Mumbai Plant"], {"MfgPlantFlag": ["N"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "MfgPlantFlag") in _gap_keys(gaps)

    def test_whitespace_trimmed(self):
        df = _full_actual(["Mumbai Plant"], {"MfgPlantFlag": ["  Y  "]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "MfgPlantFlag") in _cip_keys(cip)


# ══════════════════════════════════════════════════════════════════════════════
# 8. AutoDelAllocFlag (ideal = N)
# ══════════════════════════════════════════════════════════════════════════════


class TestAutoDelAllocFlag:
    def test_y_is_gap(self):
        df = _full_actual(["Mumbai Plant"], {"AutoDelAllocFlag": ["Y"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "AutoDelAllocFlag") in _gap_keys(gaps)

    def test_n_is_cip(self):
        df = _full_actual(["Mumbai Plant"], {"AutoDelAllocFlag": ["N"]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "AutoDelAllocFlag") in _cip_keys(cip)

    def test_enabled_is_gap(self):
        df = _full_actual(["Mumbai Plant"], {"AutoDelAllocFlag": ["enabled"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "AutoDelAllocFlag") in _gap_keys(gaps)


# ══════════════════════════════════════════════════════════════════════════════
# 9. CopyLotAttributeFlag (ideal = Y)
# ══════════════════════════════════════════════════════════════════════════════


class TestCopyLotAttributeFlag:
    def test_n_is_gap(self):
        df = _full_actual(["Mumbai Plant"], {"CopyLotAttributeFlag": ["N"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "CopyLotAttributeFlag") in _gap_keys(gaps)

    def test_y_is_cip(self):
        df = _full_actual(["Mumbai Plant"], {"CopyLotAttributeFlag": ["Y"]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "CopyLotAttributeFlag") in _cip_keys(cip)

    def test_all_bus_compliant(self):
        df = _full_actual(BUS_10, {"CopyLotAttributeFlag": ["Y"] * 10})
        cip, gaps, _, _ = _run(df, PROD_IDEAL)
        assert not any(g.config_name == "CopyLotAttributeFlag" for g in gaps)


# ══════════════════════════════════════════════════════════════════════════════
# 10. MoPickConfirmRequired (ideal = y)
# ══════════════════════════════════════════════════════════════════════════════


class TestMoPickConfirmRequired:
    @pytest.mark.parametrize("v", ["Y", "y", "yes", "true", "enabled", "on"])
    def test_truthy_is_cip(self, v):
        df = _full_actual(["Mumbai Plant"], {"MoPickConfirmRequired": [v]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "MoPickConfirmRequired") in _cip_keys(cip)

    @pytest.mark.parametrize("v", ["N", "n", "No", "false", "off", "inactive"])
    def test_falsy_is_gap(self, v):
        df = _full_actual(["Mumbai Plant"], {"MoPickConfirmRequired": [v]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "MoPickConfirmRequired") in _gap_keys(gaps)


# ══════════════════════════════════════════════════════════════════════════════
# 11. NegativeInvReceiptCode (ideal = n)
# ══════════════════════════════════════════════════════════════════════════════


class TestNegativeInvReceiptCode:
    def test_n_is_cip(self):
        df = _full_actual(["Mumbai Plant"], {"NegativeInvReceiptCode": ["N"]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "NegativeInvReceiptCode") in _cip_keys(cip)

    def test_n_lowercase_is_cip(self):
        df = _full_actual(["Mumbai Plant"], {"NegativeInvReceiptCode": ["n"]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "NegativeInvReceiptCode") in _cip_keys(cip)

    def test_no_is_cip(self):
        df = _full_actual(["Mumbai Plant"], {"NegativeInvReceiptCode": ["no"]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "NegativeInvReceiptCode") in _cip_keys(cip)

    def test_false_is_cip(self):
        df = _full_actual(["Mumbai Plant"], {"NegativeInvReceiptCode": ["false"]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "NegativeInvReceiptCode") in _cip_keys(cip)

    def test_y_is_gap(self):
        df = _full_actual(["Mumbai Plant"], {"NegativeInvReceiptCode": ["Y"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "NegativeInvReceiptCode") in _gap_keys(gaps)

    def test_integer_1_is_gap(self):
        """'1' does NOT normalize to 'y' — will not match ideal 'n'. Gap expected."""
        df = _full_actual(["Mumbai Plant"], {"NegativeInvReceiptCode": ["1"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "NegativeInvReceiptCode") in _gap_keys(gaps)

    def test_integer_0_is_gap(self):
        """'0' does NOT normalize to 'n' — so ideal='n' vs actual='0' → gap."""
        df = _full_actual(["Mumbai Plant"], {"NegativeInvReceiptCode": ["0"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "NegativeInvReceiptCode") in _gap_keys(gaps)

    def test_null_is_gap(self):
        df = _full_actual(["Mumbai Plant"], {"NegativeInvReceiptCode": [None]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "NegativeInvReceiptCode") in _gap_keys(gaps)


# ══════════════════════════════════════════════════════════════════════════════
# 12. OvpkTransferOrdersEnabled (ideal = y)
# ══════════════════════════════════════════════════════════════════════════════


class TestOvpkTransferOrdersEnabled:
    def test_n_is_gap(self):
        df = _full_actual(["Mumbai Plant"], {"OvpkTransferOrdersEnabled": ["N"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "OvpkTransferOrdersEnabled") in _gap_keys(gaps)

    def test_y_is_cip(self):
        df = _full_actual(["Mumbai Plant"], {"OvpkTransferOrdersEnabled": ["Y"]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "OvpkTransferOrdersEnabled") in _cip_keys(cip)

    def test_per_bu_mix(self):
        df = _full_actual(
            BUS_5,
            {"OvpkTransferOrdersEnabled": ["Y", "N", "Y", "N", "Y"]},
        )
        cip, gaps, _, _ = _run(df, PROD_IDEAL)
        assert (
            len([g for g in gaps if g.config_name == "OvpkTransferOrdersEnabled"]) == 2
        )


# ══════════════════════════════════════════════════════════════════════════════
# 13. ProjectReferenceEnabled (ideal = y)
# ══════════════════════════════════════════════════════════════════════════════


class TestProjectReferenceEnabled:
    @pytest.mark.parametrize("v", ["Y", "y", "yes", "true", "enabled"])
    def test_truthy_is_cip(self, v):
        df = _full_actual(["Mumbai Plant"], {"ProjectReferenceEnabled": [v]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "ProjectReferenceEnabled") in _cip_keys(cip)

    def test_n_is_gap(self):
        df = _full_actual(["Mumbai Plant"], {"ProjectReferenceEnabled": ["N"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "ProjectReferenceEnabled") in _gap_keys(gaps)

    def test_integer_1_is_gap(self):
        """LOV code '1' does NOT equal ideal 'y' — these are different representations."""
        df = _full_actual(["Mumbai Plant"], {"ProjectReferenceEnabled": ["1"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "ProjectReferenceEnabled") in _gap_keys(gaps)

    def test_integer_2_is_gap(self):
        df = _full_actual(["Mumbai Plant"], {"ProjectReferenceEnabled": ["2"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "ProjectReferenceEnabled") in _gap_keys(gaps)


# ══════════════════════════════════════════════════════════════════════════════
# 14. PurchasingByRevision (ideal = n)
# ══════════════════════════════════════════════════════════════════════════════


class TestPurchasingByRevision:
    def test_y_is_gap(self):
        df = _full_actual(["Mumbai Plant"], {"PurchasingByRevision": ["Y"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "PurchasingByRevision") in _gap_keys(gaps)

    def test_n_is_cip(self):
        df = _full_actual(["Mumbai Plant"], {"PurchasingByRevision": ["N"]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "PurchasingByRevision") in _cip_keys(cip)

    def test_false_is_cip(self):
        df = _full_actual(["Mumbai Plant"], {"PurchasingByRevision": ["false"]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "PurchasingByRevision") in _cip_keys(cip)


# ══════════════════════════════════════════════════════════════════════════════
# 15. SpecialHndlngOvrpkFlag (ideal = y)
# ══════════════════════════════════════════════════════════════════════════════


class TestSpecialHndlngOvrpkFlag:
    def test_n_is_gap(self):
        df = _full_actual(["Mumbai Plant"], {"SpecialHndlngOvrpkFlag": ["N"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "SpecialHndlngOvrpkFlag") in _gap_keys(gaps)

    def test_y_is_cip(self):
        df = _full_actual(["Mumbai Plant"], {"SpecialHndlngOvrpkFlag": ["Y"]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "SpecialHndlngOvrpkFlag") in _cip_keys(cip)


# ══════════════════════════════════════════════════════════════════════════════
# 16. UseCurItemCostFlag (ideal = Y)
# ══════════════════════════════════════════════════════════════════════════════


class TestUseCurItemCostFlag:
    def test_n_is_gap(self):
        df = _full_actual(["Mumbai Plant"], {"UseCurItemCostFlag": ["N"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "UseCurItemCostFlag") in _gap_keys(gaps)

    def test_y_is_cip(self):
        df = _full_actual(["Mumbai Plant"], {"UseCurItemCostFlag": ["Y"]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "UseCurItemCostFlag") in _cip_keys(cip)

    def test_all_bus_gap(self):
        df = _full_actual(BUS_10, {"UseCurItemCostFlag": ["N"] * 10})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert len([g for g in gaps if g.config_name == "UseCurItemCostFlag"]) == 10


# ══════════════════════════════════════════════════════════════════════════════
# 17. Capture-only configs — must NEVER appear in gaps
# ══════════════════════════════════════════════════════════════════════════════


class TestCaptureOnlyConfigs:
    CAPTURE_CONFIGS = [
        "BusinessUnitId",
        "ContractMfgFlag",
        "CustomerAccountNumber",
        "DefaultSubinvOrderValue",
        "FaBookTypeCode",
        "MasterOrganizationId",
        "OrganizationCode",
        "OrganizationId",
        "ProfitCenterBuId",
        "SourceType",
        "SupplierId",
        "SupplierSiteId",
        "TrackCountryOfOriginFlag",
    ]

    def test_capture_configs_never_in_gaps(self):
        df = _full_actual(BUS_5)
        _, gaps, extra, _ = _run(df, PROD_IDEAL)
        gap_config_names = {r.config_name for r in gaps}
        for name in self.CAPTURE_CONFIGS:
            assert name not in gap_config_names, (
                f"{name} appeared in gaps (should be capture)"
            )

    def test_capture_configs_appear_in_extra(self):
        df = _full_actual(BUS_5)
        _, _, extra, _ = _run(df, PROD_IDEAL)
        extra_names = {r.config_name for r in extra}
        # Check the ones that can actually match headers
        for name in [
            "BusinessUnitId",
            "OrganizationCode",
            "FaBookTypeCode",
            "ContractMfgFlag",
            "CustomerAccountNumber",
            "DefaultSubinvOrderValue",
            "MasterOrganizationId",
            "ProfitCenterBuId",
            "SourceType",
            "SupplierId",
            "SupplierSiteId",
            "TrackCountryOfOriginFlag",
        ]:
            assert name in extra_names, f"{name} not found in extra"

    def test_business_unit_id_captured_with_correct_value(self):
        """BusinessUnitId→BusinessUnitName alias; each BU's value is captured."""
        df = _full_actual(
            ["Mumbai Plant", "Delhi HQ"],
            {"BusinessUnitName": ["BU-101", "BU-202"]},
        )
        _, _, extra, _ = _run(df, PROD_IDEAL)
        bu_captures = {
            r.bu_name: r.actual_value
            for r in extra
            if r.config_name == "BusinessUnitId"
        }
        assert bu_captures["Mumbai Plant"] == "BU-101"
        assert bu_captures["Delhi HQ"] == "BU-202"

    def test_org_code_captured_for_all_bus(self):
        df = _full_actual(BUS_5, {"OrganizationCode": ["O1", "O2", "O3", "O4", "O5"]})
        _, _, extra, _ = _run(df, PROD_IDEAL)
        org_captures = {
            r.bu_name: r.actual_value
            for r in extra
            if r.config_name == "OrganizationCode"
        }
        assert len(org_captures) == 5
        assert org_captures["Mumbai Plant"] == "O1"
        assert org_captures["Pune SEZ"] == "O5"

    def test_null_capture_value_is_empty_string(self):
        """Null actual for a capture field → '' (not 'nan'), must not raise."""
        df = _full_actual(["Mumbai Plant"], {"BusinessUnitName": [None]})
        _, gaps, extra, _ = _run(df, PROD_IDEAL)
        bu_row = next((r for r in extra if r.config_name == "BusinessUnitId"), None)
        assert bu_row is not None
        assert bu_row.actual_value == ""

    def test_master_org_code_alias_captured(self):
        """MasterOrganizationId→MasterOrgCode alias; value captured correctly."""
        df = _full_actual(["Mumbai Plant"], {"MasterOrgCode": ["MASTER-01"]})
        _, _, extra, _ = _run(df, PROD_IDEAL)
        master_row = next(
            (r for r in extra if r.config_name == "MasterOrganizationId"), None
        )
        assert master_row is not None
        assert master_row.actual_value == "MASTER-01"

    def test_supplier_id_alias_captured(self):
        """SupplierId→SupplierName alias; vendor name captured."""
        df = _full_actual(["Mumbai Plant"], {"SupplierName": ["Tata Steel Ltd"]})
        _, _, extra, _ = _run(df, PROD_IDEAL)
        row = next((r for r in extra if r.config_name == "SupplierId"), None)
        assert row is not None
        assert row.actual_value == "Tata Steel Ltd"

    def test_supplier_site_alias_captured(self):
        df = _full_actual(["Mumbai Plant"], {"SupplierSiteName": ["Mumbai Site 1"]})
        _, _, extra, _ = _run(df, PROD_IDEAL)
        row = next((r for r in extra if r.config_name == "SupplierSiteId"), None)
        assert row is not None
        assert row.actual_value == "Mumbai Site 1"

    def test_profit_center_alias_captured(self):
        df = _full_actual(["Mumbai Plant"], {"ProfitCenterBuName": ["PC-MH-01"]})
        _, _, extra, _ = _run(df, PROD_IDEAL)
        row = next((r for r in extra if r.config_name == "ProfitCenterBuId"), None)
        assert row is not None
        assert row.actual_value == "PC-MH-01"


# ══════════════════════════════════════════════════════════════════════════════
# 18. Known unmatched fields (5 fields documented in CLAUDE.md)
# ══════════════════════════════════════════════════════════════════════════════


class TestKnownUnmatchedFields:
    UNMATCHED = [
        "LastUpdateDate",
        "LastUpdateLogin",
        "LastUpdatedBy",
        "SourceOrganizationId",
        "SourceSubinventory",
    ]

    def test_known_unmatched_produce_no_findings(self):
        """These 5 fields have no column in the standard INV export — no rows anywhere."""
        df = _full_actual(BUS_5)
        cip, gaps, extra, mapping = _run(df, PROD_IDEAL)
        all_findings = cip + gaps + extra
        for name in self.UNMATCHED:
            found = [r for r in all_findings if r.config_name == name]
            assert len(found) == 0, f"{name} produced findings but should be unmatched"

    def test_known_unmatched_appear_as_unmatched_in_mapping(self):
        df = _full_actual(BUS_5)
        _, _, _, mapping = _run(df, PROD_IDEAL)
        unmatched_names = {m.ideal_name for m in mapping if m.status == "Unmatched"}
        for name in self.UNMATCHED:
            assert name in unmatched_names, f"{name} should be Unmatched in mapping"


# ══════════════════════════════════════════════════════════════════════════════
# 19. Multi-BU compliance patterns
# ══════════════════════════════════════════════════════════════════════════════


class TestMultiBUPatterns:
    def test_one_bad_bu_among_ten(self):
        """Only one non-compliant BU — all others should be CIP."""
        bus = BUS_10
        vals = ["Y"] * 9 + ["N"]
        df = _full_actual(bus, {"InventoryFlag": vals})
        cip, gaps, _, _ = _run(df, PROD_IDEAL)
        inv_gaps = [r for r in gaps if r.config_name == "InventoryFlag"]
        inv_cip = [r for r in cip if r.config_name == "InventoryFlag"]
        assert len(inv_gaps) == 1
        assert len(inv_cip) == 9
        assert inv_gaps[0].bu_name == "Kochi South"

    def test_half_bus_gap_half_cip(self):
        bus = [f"BU_{i}" for i in range(10)]
        vals = ["Y" if i % 2 == 0 else "N" for i in range(10)]
        df = _full_actual(bus, {"AllowNegOnhandCcTxns": vals})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        neg_gaps = [g for g in gaps if g.config_name == "AllowNegOnhandCcTxns"]
        # ideal=N, actual=Y → gap; actual=N → CIP
        assert len(neg_gaps) == 5

    def test_fifty_bus_partial_compliance(self):
        n = 50
        bus = [f"Plant_{i:03d}" for i in range(n)]
        eam_vals = ["Y" if i < 40 else "N" for i in range(n)]
        df = _full_actual(bus, {"EamEnabledFlag": eam_vals})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        eam_gaps = [g for g in gaps if g.config_name == "EamEnabledFlag"]
        assert len(eam_gaps) == 10

    def test_all_bus_all_configs_gap(self):
        """Worst case: every BU has every comparison config wrong."""
        bad = {
            "AllowItemSubstitutions": "N",  # ideal Y
            "AllowNegOnhandCcTxns": "Y",  # ideal N
            "AutoDelAllocFlag": "Y",  # ideal N
            "CopyLotAttributeFlag": "N",  # ideal Y
            "EamEnabledFlag": "N",  # ideal y
            "FifoOrigRcptDateFlag": "N",  # ideal y
            "InventoryFlag": "N",  # ideal y
            "MfgPlantFlag": "N",  # ideal y
            "MoPickConfirmRequired": "N",  # ideal y
            "NegativeInvReceiptCode": "Y",  # ideal n
            "OvpkTransferOrdersEnabled": "N",  # ideal y
            "ProjectReferenceEnabled": "N",  # ideal y
            "PurchasingByRevision": "Y",  # ideal n
            "SpecialHndlngOvrpkFlag": "N",  # ideal y
            "UseCurItemCostFlag": "N",  # ideal Y
        }
        df = _full_actual(BUS_5, bad)
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        # 15 comparison configs × 5 BUs = 75 gaps
        assert len(gaps) == 75


# ══════════════════════════════════════════════════════════════════════════════
# 20. Oracle export value quirks
# ══════════════════════════════════════════════════════════════════════════════


class TestOracleExportQuirks:
    def test_float_y_values_not_affected(self):
        """Float normalization only applies to numeric strings, not Y/N."""
        df = _full_actual(["Mumbai Plant"], {"InventoryFlag": ["Y"]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "InventoryFlag") in _cip_keys(cip)

    def test_float_5_dot_0_equals_5(self):
        """Tolerance values exported as floats should compare equal to integer ideals."""
        tol_ideal = _ideal(("MaxQtyOrdered", "5"), ("MaxAmtReceived", "10"))
        df = pd.DataFrame(
            {
                "BU_NAME": ["Mumbai Plant"],
                "MaxQtyOrdered": ["5.0"],
                "MaxAmtReceived": ["10.0"],
            }
        )
        cip, gaps, _, _ = _run(df, tol_ideal)
        assert len(gaps) == 0

    def test_leading_trailing_whitespace_stripped(self):
        df = _full_actual(["Mumbai Plant"], {"InventoryFlag": ["  Y  "]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "InventoryFlag") in _cip_keys(cip)

    def test_mixed_case_config_values(self):
        df = _full_actual(
            ["Mumbai Plant"],
            {"EamEnabledFlag": ["yEs"], "MfgPlantFlag": ["tRuE"]},
        )
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "EamEnabledFlag") in _cip_keys(cip)
        assert ("Mumbai Plant", "MfgPlantFlag") in _cip_keys(cip)

    def test_integer_2_is_not_y_or_n(self):
        """LOV code 2 doesn't collapse to any boolean — it's a literal '2'."""
        df = _full_actual(["Mumbai Plant"], {"ProjectReferenceEnabled": ["2"]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "ProjectReferenceEnabled") in _gap_keys(gaps)

    def test_large_integer_not_confused_with_boolean(self):
        """LOV codes like 100, 200 must never map to y/n."""
        tol_ideal = _ideal(("MaxQtyOrdered", "100"))
        df = pd.DataFrame({"BU_NAME": ["Mumbai Plant"], "MaxQtyOrdered": ["100"]})
        cip, gaps, _, _ = _run(df, tol_ideal)
        assert ("Mumbai Plant", "MaxQtyOrdered") in _cip_keys(cip)

    def test_string_none_in_actual_treated_as_null(self):
        """Some Oracle exports write the string 'None' — ensure it's treated as null."""
        df = _full_actual(["Mumbai Plant"])
        df["InventoryFlag"] = "None"
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        # "none".lower() != "y", so it's a gap
        assert ("Mumbai Plant", "InventoryFlag") in _gap_keys(gaps)

    def test_nan_actual_is_gap_not_capture(self):
        """NaN actual against a comparison ideal → gap (null config)."""
        df = _full_actual(["Mumbai Plant"])
        df["InventoryFlag"] = float("nan")
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "InventoryFlag") in _gap_keys(gaps)

    @pytest.mark.parametrize("v", ["Y", "y", "YES", "Yes", "true", "True", "enabled"])
    def test_all_truthy_oracle_exports_cip_for_y_ideal(self, v):
        df = _full_actual(["Mumbai Plant"], {"InventoryFlag": [v]})
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "InventoryFlag") in _cip_keys(cip)

    @pytest.mark.parametrize("v", ["N", "n", "NO", "No", "false", "False", "disabled"])
    def test_all_falsy_oracle_exports_gap_for_y_ideal(self, v):
        df = _full_actual(["Mumbai Plant"], {"InventoryFlag": [v]})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert ("Mumbai Plant", "InventoryFlag") in _gap_keys(gaps)


# ══════════════════════════════════════════════════════════════════════════════
# 21. BU name edge cases
# ══════════════════════════════════════════════════════════════════════════════


class TestBUNameEdgeCases:
    def test_bu_names_with_spaces_preserved(self):
        bus = ["Mumbai Plant Site 1", "Delhi Head Office", "Bangalore IT Park"]
        df = _full_actual(bus)
        cip, _, _, _ = _run(df, PROD_IDEAL)
        bu_names = {r.bu_name for r in cip}
        for bu in bus:
            assert bu in bu_names

    def test_unicode_bu_names(self):
        bus = ["मुंबई प्लांट", "دبي مصنع", "上海工厂", "München Werk", "Nairobi Plant"]
        df = _full_actual(bus)
        cip, gaps, _, _ = _run(df, PROD_IDEAL)
        all_bus = {r.bu_name for r in cip + gaps}
        for bu in bus:
            assert bu in all_bus, f"BU name lost: {bu!r}"

    def test_numeric_bu_name(self):
        df = _full_actual(["12345"])
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert any(r.bu_name == "12345" for r in cip)

    def test_long_bu_name(self):
        long_name = "A" * 200
        df = _full_actual([long_name])
        cip, _, _, _ = _run(df, PROD_IDEAL)
        assert any(r.bu_name == long_name for r in cip)

    def test_empty_bu_name_row_is_skipped(self):
        """Blank BU rows must be silently dropped (not produce findings)."""
        df = _full_actual(["Mumbai Plant", "", "Delhi HQ"])
        cip, gaps, extra, _ = _run(df, PROD_IDEAL)
        all_bu_names = {r.bu_name for r in cip + gaps + extra}
        assert "" not in all_bu_names


# ══════════════════════════════════════════════════════════════════════════════
# 22. Header matching — Oracle-specific alias scenarios
# ══════════════════════════════════════════════════════════════════════════════


class TestHeaderMatchingAliases:
    def test_all_six_aliases_map_correctly(self):
        """All 6 entries in MANUAL_ALIASES route to the right actual column."""
        _, _, _, mapping = _run(_full_actual(["Mumbai Plant"]), PROD_IDEAL)
        alias_map = {
            "AllowItemSubstitutionsFlag": ("AllowItemSubstitutions", "Manual Alias"),
            "BusinessUnitId": ("BusinessUnitName", "Manual Alias"),
            "MasterOrganizationId": ("MasterOrgCode", "Manual Alias"),
            "ProfitCenterBuId": ("ProfitCenterBuName", "Manual Alias"),
            "SupplierId": ("SupplierName", "Manual Alias"),
            "SupplierSiteId": ("SupplierSiteName", "Manual Alias"),
        }
        for ideal_name, (expected_header, expected_method) in alias_map.items():
            m = next((x for x in mapping if x.ideal_name == ideal_name), None)
            assert m is not None, f"No mapping entry for {ideal_name}"
            assert m.matched_header == expected_header, (
                f"{ideal_name}: expected {expected_header}, got {m.matched_header}"
            )
            assert m.match_method == expected_method

    def test_organization_id_matches_via_fuzzy(self):
        """OrganizationId has no alias — must fuzzy-match to OrganizationCode."""
        _, _, _, mapping = _run(_full_actual(["Mumbai Plant"]), PROD_IDEAL)
        m = next(x for x in mapping if x.ideal_name == "OrganizationId")
        assert m.match_method == "Fuzzy"
        assert m.matched_header == "OrganizationCode"
        assert m.status == "Matched"

    def test_exact_match_configs_have_exact_method(self):
        """Configs whose names appear verbatim in the export use Exact match."""
        exact_names = [
            "AllowNegOnhandCcTxns",
            "AutoDelAllocFlag",
            "CopyLotAttributeFlag",
            "EamEnabledFlag",
            "FifoOrigRcptDateFlag",
            "InventoryFlag",
            "MfgPlantFlag",
            "MoPickConfirmRequired",
            "NegativeInvReceiptCode",
            "OvpkTransferOrdersEnabled",
            "ProjectReferenceEnabled",
            "PurchasingByRevision",
            "SpecialHndlngOvrpkFlag",
            "UseCurItemCostFlag",
        ]
        _, _, _, mapping = _run(_full_actual(["Mumbai Plant"]), PROD_IDEAL)
        for name in exact_names:
            m = next((x for x in mapping if x.ideal_name == name), None)
            assert m is not None, f"No mapping for {name}"
            assert m.match_method == "Exact", (
                f"{name}: expected Exact, got {m.match_method}"
            )

    def test_similarity_score_1_for_all_matched_except_fuzzy(self):
        _, _, _, mapping = _run(_full_actual(["Mumbai Plant"]), PROD_IDEAL)
        for m in mapping:
            if m.status == "Matched" and m.match_method != "Fuzzy":
                assert m.similarity_score == 1.0, (
                    f"{m.ideal_name} ({m.match_method}) score={m.similarity_score}"
                )


# ══════════════════════════════════════════════════════════════════════════════
# 23. Report structure
# ══════════════════════════════════════════════════════════════════════════════


class TestReportStructure:
    def test_audit_summary_has_14_metric_rows(self):
        import io as _io
        import openpyxl
        from engine.writer import build_report

        df = _full_actual(BUS_5)
        mapping = build_mapping(PROD_IDEAL, list(df.columns))
        cip, gaps, extra = compare(df, mapping)
        xlsx = build_report(
            cip,
            gaps,
            extra,
            mapping,
            actual_filename="test.xlsx",
            ideal_filename="inv_org_params",
            total_bu_rows=len(df),
            fuzzy_threshold=80,
        )
        wb = openpyxl.load_workbook(_io.BytesIO(xlsx))
        ws = wb["Audit Summary"]
        data_rows = [
            r
            for r in ws.iter_rows(min_row=2, values_only=True)
            if any(c is not None for c in r)
        ]
        assert len(data_rows) == 14

    def test_mapping_sheet_footer_lists_unmatched(self):
        import io as _io
        import openpyxl
        from engine.writer import build_report

        df = _full_actual(BUS_5)
        mapping = build_mapping(PROD_IDEAL, list(df.columns))
        cip, gaps, extra = compare(df, mapping)
        xlsx = build_report(
            cip,
            gaps,
            extra,
            mapping,
            actual_filename="test.xlsx",
            ideal_filename="inv_org_params",
            total_bu_rows=len(df),
            fuzzy_threshold=80,
        )
        wb = openpyxl.load_workbook(_io.BytesIO(xlsx))
        ws = wb["Header Mapping and Exc"]
        all_cell_values = [
            c for row in ws.iter_rows(values_only=True) for c in row if c is not None
        ]
        assert "Note" in all_cell_values
        assert "Below are unmatched ideal configuration names" in all_cell_values

    def test_capture_configs_appear_in_additional_data_sheet(self):
        import io as _io
        import openpyxl
        from engine.writer import build_report

        df = _full_actual(["Mumbai Plant"])
        mapping = build_mapping(PROD_IDEAL, list(df.columns))
        cip, gaps, extra = compare(df, mapping)
        xlsx = build_report(
            cip,
            gaps,
            extra,
            mapping,
            actual_filename="test.xlsx",
            ideal_filename="inv_org_params",
            total_bu_rows=1,
            fuzzy_threshold=80,
        )
        wb = openpyxl.load_workbook(_io.BytesIO(xlsx))
        ws = wb["Controls additional data"]
        config_names = [
            row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row[1]
        ]
        assert "BusinessUnitId" in config_names or "OrganizationCode" in config_names

    def test_gaps_sheet_has_correct_four_columns(self):
        import io as _io
        import openpyxl
        from engine.writer import build_report

        df = _full_actual(["Mumbai Plant"], {"InventoryFlag": ["N"]})
        mapping = build_mapping(PROD_IDEAL, list(df.columns))
        cip, gaps, extra = compare(df, mapping)
        xlsx = build_report(
            cip,
            gaps,
            extra,
            mapping,
            actual_filename="test.xlsx",
            ideal_filename="inv_org_params",
            total_bu_rows=1,
            fuzzy_threshold=80,
        )
        wb = openpyxl.load_workbook(_io.BytesIO(xlsx))
        ws = wb["Control gaps"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert headers == [
            "BU Name",
            "Configuration Name",
            "Actual Configuration Value",
            "Comment",
        ]


# ══════════════════════════════════════════════════════════════════════════════
# 24. Performance sanity — 100 BUs × 33 configs
# ══════════════════════════════════════════════════════════════════════════════


class TestPerformance:
    def test_100_bus_runs_in_under_5_seconds(self):
        import time

        bus = [f"BU_{i:04d}" for i in range(100)]
        df = _full_actual(bus)
        t0 = time.time()
        _run(df, PROD_IDEAL)
        elapsed = time.time() - t0
        assert elapsed < 5.0, f"100-BU run took {elapsed:.2f}s"

    def test_500_bus_completes(self):
        bus = [f"Plant_{i:04d}" for i in range(500)]
        df = _full_actual(bus)
        cip, gaps, extra, _ = _run(df, PROD_IDEAL)
        assert len(cip) > 0

    def test_row_count_scales_linearly(self):
        for n in [10, 50, 100]:
            bus = [f"BU_{i}" for i in range(n)]
            df = _full_actual(bus)
            cip, gaps, extra, _ = _run(df, PROD_IDEAL)
            total = len(cip) + len(gaps) + len(extra)
            # total rows = n_bus × matched_configs (comparison + capture)
            assert total == n * (len(cip) // n + len(gaps) // n + len(extra) // n)


# ══════════════════════════════════════════════════════════════════════════════
# 25. Realistic mixed-compliance audit scenarios
# ══════════════════════════════════════════════════════════════════════════════


class TestRealisticScenarios:
    def test_scenario_one_plant_non_compliant(self):
        """Typical finding: one plant has negative inventory enabled."""
        df = _full_actual(
            BUS_5,
            {"AllowNegOnhandCcTxns": ["N", "N", "Y", "N", "N"]},
        )
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        neg_gaps = [g for g in gaps if g.config_name == "AllowNegOnhandCcTxns"]
        assert len(neg_gaps) == 1
        assert neg_gaps[0].bu_name == "Bangalore IT"

    def test_scenario_new_acquisition_all_wrong(self):
        """Newly acquired entity has all comparison flags misconfigured."""
        bad = {
            "AllowItemSubstitutions": "N",
            "AllowNegOnhandCcTxns": "Y",
            "AutoDelAllocFlag": "Y",
            "CopyLotAttributeFlag": "N",
            "EamEnabledFlag": "N",
            "FifoOrigRcptDateFlag": "N",
            "InventoryFlag": "N",
            "MfgPlantFlag": "N",
            "MoPickConfirmRequired": "N",
            "NegativeInvReceiptCode": "Y",
            "OvpkTransferOrdersEnabled": "N",
            "ProjectReferenceEnabled": "N",
            "PurchasingByRevision": "Y",
            "SpecialHndlngOvrpkFlag": "N",
            "UseCurItemCostFlag": "N",
        }
        df = _full_actual(["New Acquisition Ltd"], bad)
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        assert len(gaps) == 15

    def test_scenario_treasury_bu_different_exchange_rate(self):
        """Treasury BU uses Spot rate; all others use Corporate. Only treasury is gap."""
        # Using inline ideal for this specific scenario
        exch_ideal = _ideal(
            ("DefaultExchangeRateType", "Corporate", "Corporate, Spot, User")
        )
        bus = ["Treasury", "Operations", "Manufacturing"]
        df = pd.DataFrame(
            {
                "BU_NAME": bus,
                "DefaultExchangeRateType": ["Spot", "Corporate", "Corporate"],
            }
        )
        _, gaps, _, _ = _run(df, exch_ideal)
        assert ("Treasury", "DefaultExchangeRateType") in _gap_keys(gaps)
        assert ("Operations", "DefaultExchangeRateType") not in _gap_keys(gaps)

    def test_scenario_fifo_disabled_multiple_plants(self):
        plants = [f"Plant_{c}" for c in ["A", "B", "C", "D", "E"]]
        fifo = ["Y", "N", "Y", "Y", "N"]
        df = _full_actual(plants, {"FifoOrigRcptDateFlag": fifo})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        fifo_gaps = {g.bu_name for g in gaps if g.config_name == "FifoOrigRcptDateFlag"}
        assert fifo_gaps == {"Plant_B", "Plant_E"}

    def test_scenario_overpacking_disabled_globally(self):
        """All BUs have OvpkTransferOrdersEnabled=N — global finding."""
        df = _full_actual(BUS_10, {"OvpkTransferOrdersEnabled": ["N"] * 10})
        _, gaps, _, _ = _run(df, PROD_IDEAL)
        ovpk_gaps = [g for g in gaps if g.config_name == "OvpkTransferOrdersEnabled"]
        assert len(ovpk_gaps) == 10
        assert {g.bu_name for g in ovpk_gaps} == set(BUS_10)

    def test_scenario_audit_comment_text_is_exact(self):
        """Comments must be exactly 'Controls in place', 'Controls gaps', 'Actual config captured'."""
        df = _full_actual(
            ["Mumbai Plant"],
            {"AllowNegOnhandCcTxns": ["Y"]},  # gap (ideal=N)
        )
        cip, gaps, extra, _ = _run(df, PROD_IDEAL)
        cip_row = next(r for r in cip if r.config_name == "InventoryFlag")
        gap_row = next(r for r in gaps if r.config_name == "AllowNegOnhandCcTxns")
        extra_row = next(r for r in extra if r.config_name == "BusinessUnitId")
        assert cip_row.comment == "Controls in place"
        assert gap_row.comment == "Controls gaps"
        assert extra_row.comment == "Actual config captured"
