"""Integration test — full pipeline with realistic Oracle Fusion INV data.

BU names, config values, and org codes mirror real client export patterns.
No actual PII is used.
"""

import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import io
import pandas as pd
import openpyxl
from engine import build_mapping, build_report, compare, read_ideal


# ── realistic client data ─────────────────────────────────────────────────────

ORACLE_BUS = [
    "Mumbai Plant",
    "Delhi HQ",
    "Bangalore IT",
    "Singapore Regional",
    "UAE Free Zone",
    "London EMEA",
    "New York Americas",
    "Tokyo APAC",
    "Sydney Oceania",
    "Cape Town Africa",
]

ORACLE_ACTUAL = {
    # Comparison configs — some pass, some fail
    "AllowItemSubstitutions": ["Y", "N", "Y", "Y", "Y", "Y", "N", "Y", "Y", "Y"],
    "AllowNegOnhandCcTxns": ["N", "N", "Y", "N", "N", "N", "N", "N", "N", "N"],
    "EamEnabledFlag": ["Y", "y", "N", None, "Y", "Y", "Y", "Y", "N", "Y"],
    "FifoOrigRcptDateFlag": ["Y", "Y", "Y", "Y", "N", "Y", "Y", "Y", "Y", "Y"],
    "InventoryFlag": ["Y", "Y", "Y", "Y", "Y", "Y", "Y", "Y", "Y", "Y"],
    "MfgPlantFlag": ["Y", "Y", None, "Y", "Y", "Y", "Y", "Y", None, "Y"],
    "FABookTypeCode": ["N", "N", "N", "N", "N", "N", "N", "N", "N", "N"],
    # Capture-only — no comparison
    "BusinessUnitName": [
        "Mumbai BU",
        "Delhi BU",
        "Blr BU",
        "SG BU",
        "UAE BU",
        "London BU",
        "NY BU",
        "Tokyo BU",
        "Sydney BU",
        "CT BU",
    ],
    "OrganizationCode": [2, 50, 51, 200, 300, 400, 500, 600, 700, 800],
    "MasterOrgCode": [0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    "ProfitCenterBuName": [
        "Mumbai BU",
        "Delhi BU",
        "Blr BU",
        "SG BU",
        "UAE BU",
        "London BU",
        "NY BU",
        "Tokyo BU",
        "Sydney BU",
        "CT BU",
    ],
}

ORACLE_IDEAL = [
    # Compared
    ("AllowItemSubstitutionsFlag", "Y"),
    ("AllowNegOnhandCcTxns", "N"),
    ("EamEnabledFlag", "y"),
    ("FifoOrigRcptDateFlag", "y"),
    ("InventoryFlag", "y"),
    ("MfgPlantFlag", "y"),
    ("FaBookTypeCode", "N"),
    # Captured
    ("BusinessUnitId", "Capture the value defined"),
    ("OrganizationId", "Capture the value defined"),
    ("MasterOrganizationId", "Capture the value defined"),
    ("ProfitCenterBuId", "Capture the value defined"),
]


def make_actual_df() -> pd.DataFrame:
    df = pd.DataFrame({"BU_NAME": ORACLE_BUS, **ORACLE_ACTUAL})
    return df


def make_ideal_df() -> pd.DataFrame:
    return pd.DataFrame(ORACLE_IDEAL, columns=["config_name", "ideal_value"])


# ── tests ─────────────────────────────────────────────────────────────────────


class TestPipelineOutput:
    def setup_method(self):
        actual_df = make_actual_df()
        ideal_df = make_ideal_df()
        mapping = build_mapping(ideal_df, list(actual_df.columns))
        self.cip, self.gaps, self.extra = compare(actual_df, mapping)
        self.mapping = mapping

    def test_total_rows_equals_bu_count_times_matched_configs(self):
        """Every BU × every matched config must produce exactly one row."""
        matched_compared = sum(
            1
            for m in self.mapping
            if m.status == "Matched" and "capture" not in m.ideal_value.lower()
        )
        matched_capture = sum(
            1
            for m in self.mapping
            if m.status == "Matched" and "capture" in m.ideal_value.lower()
        )
        total = len(self.cip) + len(self.gaps) + len(self.extra)
        assert total == len(ORACLE_BUS) * (matched_compared + matched_capture)

    def test_all_7_compared_configs_are_matched(self):
        compared_names = {r.config_name for r in self.cip + self.gaps}
        # These 7 configs should all appear in comparison output
        expected = {
            "AllowItemSubstitutionsFlag",
            "AllowNegOnhandCcTxns",
            "EamEnabledFlag",
            "FifoOrigRcptDateFlag",
            "InventoryFlag",
            "MfgPlantFlag",
            "FaBookTypeCode",
        }
        assert expected <= compared_names

    def test_all_4_capture_configs_in_extra(self):
        captured_names = {r.config_name for r in self.extra}
        expected = {
            "BusinessUnitId",
            "OrganizationId",
            "MasterOrganizationId",
            "ProfitCenterBuId",
        }
        assert expected <= captured_names

    def test_known_gaps_are_detected(self):
        """Delhi HQ has AllowItemSubstitutions=N, ideal=Y → should be a gap."""
        gap_keys = {(r.bu_name, r.config_name) for r in self.gaps}
        assert ("Delhi HQ", "AllowItemSubstitutionsFlag") in gap_keys

    def test_known_controls_in_place(self):
        """InventoryFlag = Y for all 10 BUs, ideal = Y → all in place."""
        inv_pass = [r for r in self.cip if r.config_name == "InventoryFlag"]
        assert len(inv_pass) == 10

    def test_null_mfgplantflag_is_gap(self):
        """Bangalore IT and Sydney Oceania have null MfgPlantFlag → gaps."""
        gap_keys = {(r.bu_name, r.config_name) for r in self.gaps}
        assert ("Bangalore IT", "MfgPlantFlag") in gap_keys
        assert ("Sydney Oceania", "MfgPlantFlag") in gap_keys

    def test_lowercase_y_passes(self):
        """Delhi HQ has EamEnabledFlag='y', ideal='y' → controls in place."""
        cip_keys = {(r.bu_name, r.config_name) for r in self.cip}
        assert ("Delhi HQ", "EamEnabledFlag") in cip_keys


class TestExcelOutput:
    def setup_method(self):
        actual_df = make_actual_df()
        ideal_df = make_ideal_df()
        mapping = build_mapping(ideal_df, list(actual_df.columns))
        cip, gaps, extra = compare(actual_df, mapping)
        excel_bytes = build_report(
            cip,
            gaps,
            extra,
            mapping,
            actual_filename="oracle_inv.xlsx",
            ideal_filename="ideal_values.xlsx",
            total_bu_rows=len(actual_df),
            fuzzy_threshold=80,
        )
        self.wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True)

    def test_five_sheets_present(self):
        assert set(self.wb.sheetnames) == {
            "Controls in place",
            "Control gaps",
            "Controls additional data",
            "Header Mapping and Exc",
            "Audit Summary",
        }

    def test_data_sheet_headers(self):
        base = ["BU Name", "Configuration Name", "Actual Configuration Value"]
        for sheet_name in ["Controls in place", "Controls additional data"]:
            ws = self.wb[sheet_name]
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            assert headers == base + ["Comment"]

        # Control gaps sheet has two extra columns for auditor context
        ws = self.wb["Control gaps"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert headers == base + ["Ideal Value", "Valid Options", "Comment"]

    def test_audit_summary_has_14_rows(self):
        ws = self.wb["Audit Summary"]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 14

    def test_mapping_sheet_has_correct_columns(self):
        ws = self.wb["Header Mapping and Exc"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "Ideal Configuration Name" in headers
        assert "Mapping Type" in headers
        assert "Status" in headers


class TestReaderWriter:
    def test_read_ideal_from_csv(self):
        csv_content = "Name,Ideal value\nInventoryFlag,Y\nAllowNegOnhandCcTxns,N\n"
        df = read_ideal(io.BytesIO(csv_content.encode()))
        assert len(df) == 2
        assert list(df.columns) == ["config_name", "ideal_value", "options"]
        assert df.iloc[0]["config_name"] == "InventoryFlag"

    def test_read_ideal_strips_whitespace(self):
        csv_content = "Name,Ideal value\n  InventoryFlag  ,  Y  \n"
        df = read_ideal(io.BytesIO(csv_content.encode()))
        assert df.iloc[0]["config_name"] == "InventoryFlag"
        assert df.iloc[0]["ideal_value"] == "Y"

    def test_read_ideal_drops_empty_rows(self):
        csv_content = "Name,Ideal value\nInventoryFlag,Y\n,\n\nAllowNeg,N\n"
        df = read_ideal(io.BytesIO(csv_content.encode()))
        assert len(df) == 2
