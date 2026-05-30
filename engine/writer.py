"""Excel report builder — produces a 5-sheet styled workbook in memory.

Public API:
    build_report(...) -> bytes
"""

import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .styles import (
    BLUE_FILL,
    GREEN_FILL,
    HEADER_FILL,
    HEADER_FONT,
    RED_FILL,
    YELLOW_FILL,
    CENTER,
)
from .types import AuditRow, MappingResult


def build_report(
    controls_in_place: list[AuditRow],
    control_gaps: list[AuditRow],
    additional_data: list[AuditRow],
    mapping: list[MappingResult],
    actual_filename: str = "",
    ideal_filename: str = "",
    total_bu_rows: int = 0,
    fuzzy_threshold: int = 80,
) -> bytes:
    wb = Workbook()

    _data_sheet(wb.active, "Controls in place", controls_in_place, GREEN_FILL)
    _data_sheet(wb.create_sheet("Control gaps"), "Control gaps", control_gaps, RED_FILL)
    _data_sheet(
        wb.create_sheet("Controls additional data"),
        "Controls additional data",
        additional_data,
        BLUE_FILL,
    )
    _mapping_sheet(wb.create_sheet("Header Mapping and Exc"), mapping)
    _summary_sheet(
        wb.create_sheet("Audit Summary"),
        controls_in_place,
        control_gaps,
        additional_data,
        mapping,
        actual_filename,
        ideal_filename,
        total_bu_rows,
        fuzzy_threshold,
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _data_sheet(
    ws, title: str, rows: list[AuditRow], fill, show_options: bool = False
) -> None:
    ws.title = title
    if show_options:
        _header_row(
            ws,
            [
                "BU Name",
                "Configuration Name",
                "Actual Configuration Value",
                "Ideal Value",
                "Valid Options",
                "Comment",
            ],
        )
        for r in rows:
            ws.append(
                [
                    r.bu_name,
                    r.config_name,
                    r.actual_value,
                    r.ideal_value,
                    r.options,
                    r.comment,
                ]
            )
            _fill_row(ws, fill)
    else:
        _header_row(
            ws,
            ["BU Name", "Configuration Name", "Actual Configuration Value", "Comment"],
        )
        for r in rows:
            ws.append([r.bu_name, r.config_name, r.actual_value, r.comment])
            _fill_row(ws, fill)
    _autofit(ws)


def _mapping_sheet(ws, mapping: list[MappingResult]) -> None:
    ws.title = "Header Mapping and Exc"
    _header_row(
        ws,
        [
            "Ideal Configuration Name",
            "Actual Header Mapped",
            "Mapping Type",
            "Similarity Score",
            "Status",
            "Remarks",
        ],
    )
    for m in mapping:
        score = (
            int(m.similarity_score)
            if m.similarity_score == 1.0
            else round(m.similarity_score, 4)
        )
        ws.append(
            [
                m.ideal_name,
                m.matched_header or None,
                m.match_method,
                score,
                m.status,
                m.remarks or None,
            ]
        )
        if m.status == "Unmatched":
            _fill_row(ws, YELLOW_FILL)

    # Footer: repeat unmatched section (matches original output format)
    unmatched = [m for m in mapping if m.status == "Unmatched"]
    if unmatched:
        ws.append([None, None, None, None, None, None])
        ws.append([None, None, None, None, None, None])
        ws.append(["Note", None, None, None, None, None])
        ws.append(
            [
                "Below are unmatched ideal configuration names",
                None,
                None,
                None,
                None,
                None,
            ]
        )
        ws.append(
            [
                "Ideal Configuration Name",
                "Actual Header Mapped",
                "Mapping Type",
                "Similarity Score",
                "Status",
                "Remarks",
            ]
        )
        for m in unmatched:
            score = round(m.similarity_score, 4)
            ws.append(
                [m.ideal_name, None, m.match_method, score, m.status, m.remarks or None]
            )

    _autofit(ws)


def _summary_sheet(
    ws,
    cip,
    gaps,
    extra,
    mapping,
    actual_fn,
    ideal_fn,
    bu_rows,
    threshold,
) -> None:
    ws.title = "Audit Summary"
    _header_row(ws, ["Metric", "Value"])

    matched = sum(1 for m in mapping if m.status == "Matched")
    unmatched = sum(1 for m in mapping if m.status == "Unmatched")

    rows = [
        ("Actual file", actual_fn),
        ("Ideal file", ideal_fn),
        ("Actual sheet used", "INV_ORGANIZATION_PARAMETER"),
        ("BU column used", "Name"),
        ("Total BU rows processed", bu_rows),
        ("Total ideal configurations", len(mapping)),
        ("Matched configurations", matched),
        ("Unmatched configurations", unmatched),
        ("Compared configurations count", len({r.config_name for r in cip + gaps})),
        ("Capture-only configurations count", len({r.config_name for r in extra})),
        ("Controls in place row count", len(cip)),
        ("Control gaps row count", len(gaps)),
        ("Additional data row count", len(extra)),
        ("Fuzzy threshold used", threshold / 100),
    ]
    for metric, value in rows:
        ws.append([metric, value])

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22


# --- helpers ---


def _header_row(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 18


def _fill_row(ws, fill) -> None:
    for cell in ws[ws.max_row]:
        cell.fill = fill


def _autofit(ws) -> None:
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        width = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 4, 55)
