"""Excel style constants — one place to change colours or fonts."""

from openpyxl.styles import Alignment, Font, PatternFill

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")  # Controls in place
RED_FILL = PatternFill("solid", fgColor="FFC7CE")  # Control gaps
BLUE_FILL = PatternFill("solid", fgColor="DDEBF7")  # Additional data
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")  # Unmatched header rows

CENTER = Alignment(horizontal="center", vertical="center")
